from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import range_boundaries, get_column_letter
from dto.chart_data import ChartData, ChartSeries, DataRange
from dto.coordinate import BoundingBox
from typing import Any, List, Optional


"""
Extractor that extracts all charts from a worksheet.
Resolves cell references against the live workbook so that values and
labels are always populated — even when the chart XML contains no cache.
"""


class ChartExtractor:

    # ---- title helpers --------------------------------------------------------

    def _extract_title(self, title_obj) -> Optional[str]:
        """Extract a plain-text string from an openpyxl Title object."""
        if title_obj is None:
            return None
        try:
            tx = getattr(title_obj, "tx", None)
            rich = getattr(tx, "rich", None)
            if rich and rich.paragraphs:
                parts: List[str] = []
                for para in rich.paragraphs:
                    for run in para.r or []:
                        if run.t:
                            parts.append(run.t)
                if parts:
                    return " ".join(parts)
        except Exception:
            pass
        return None

    def _extract_axis_title(self, chart, axis_attr: str) -> Optional[str]:
        """Extract the title string from an axis (x_axis / y_axis)."""
        axis = getattr(chart, axis_attr, None)
        if axis is None:
            return None
        return self._extract_title(getattr(axis, "title", None))

    # ---- chart type -----------------------------------------------------------

    def _extract_chart_type(self, chart) -> str:
        """Derive a human-readable chart type from the openpyxl class name."""
        cls_name = type(chart).__name__  # e.g. "BarChart", "LineChart3D"
        name = cls_name.replace("Chart3D", "").replace("Chart", "").lower()
        if "3d" in cls_name.lower():
            name += "_3d"
        return name or "unknown"

    # ---- bounding box ---------------------------------------------------------

    @staticmethod
    def _marker_to_coordinate(marker) -> str:
        """
        Convert an AnchorMarker (0-indexed col/row) to a Coordinate.
        openpyxl markers are 0-based, so we add 1 for the human-readable form.
        """
        col = int(getattr(marker, "col", 0)) + 1
        row = int(getattr(marker, "row", 0)) + 1

        return f"{get_column_letter(col)}{row}"

    # Approximate default cell dimensions in EMUs (English Metric Units).
    # Excel's default row height is 15pt ≈ 0.53cm ≈ 190500 EMU.
    # Excel's default column width (8 chars + padding) ≈ 64px ≈ 609600 EMU.
    _DEFAULT_COL_EMU = 609600
    _DEFAULT_ROW_EMU = 190500

    def _bottom_right_from_extent(self, from_marker, ext) -> str:
        """
        Compute the bottom-right Coordinate for a OneCellAnchor by adding
        the extent (cx/cy in EMUs) to the starting marker position.
        """
        start_col = int(getattr(from_marker, "col", 0))  # 0-indexed
        start_row = int(getattr(from_marker, "row", 0))  # 0-indexed
        width_emu = int(getattr(ext, "cx", 0) or 0)
        height_emu = int(getattr(ext, "cy", 0) or 0)

        end_col = start_col + max(1, width_emu // self._DEFAULT_COL_EMU)
        end_row = start_row + max(1, height_emu // self._DEFAULT_ROW_EMU)

        return f"{get_column_letter(end_col + 1)}{end_row + 1}"

    def _extract_bounding_box(self, chart) -> BoundingBox:
        """
        Build a BoundingBox from the chart's drawing anchor.

        When read from a file the anchor is a TwoCellAnchor (with _from / to)
        or a OneCellAnchor (with _from and ext).  When the chart was created
        in code the anchor may still be a plain cell-reference string like "E15".
        """
        anchor = getattr(chart, "anchor", None)

        from_marker = getattr(anchor, "_from", None)
        to_marker = getattr(anchor, "to", None)
        ext = getattr(anchor, "ext", None)

        if from_marker is not None:
            top_left = self._marker_to_coordinate(from_marker)

            if to_marker is not None:
                # TwoCellAnchor — explicit bottom-right corner
                bottom_right = self._marker_to_coordinate(to_marker)
            elif ext is not None:
                # OneCellAnchor — compute bottom-right from width/height
                bottom_right = self._bottom_right_from_extent(from_marker, ext)
            else:
                bottom_right = top_left

            return BoundingBox(top_left=top_left, bottom_right=bottom_right)

        # Fallback: anchor is a plain string like "E15" (newly created chart)
        if isinstance(anchor, str) and anchor:
            col_str = "".join(c for c in anchor if c.isalpha())
            row_str = "".join(c for c in anchor if c.isdigit())
            coord = Coordinate(column=col_str, row=row_str)
            return BoundingBox(top_left=coord, bottom_right=coord)

        # No usable anchor information at all
        return BoundingBox(
            top_left=Coordinate(column="", row=""),
            bottom_right=Coordinate(column="", row=""),
        )

    # ---- ref / formula helpers ------------------------------------------------

    @staticmethod
    def _ref_formula(data_source) -> Optional[str]:
        """
        Return the range formula string (e.g. "'Sheet1'!$B$2:$B$10") from a
        val/cat data-source object, or None if unavailable.
        """
        num_ref = getattr(data_source, "numRef", None)
        if num_ref and getattr(num_ref, "f", None):
            return num_ref.f
        str_ref = getattr(data_source, "strRef", None)
        if str_ref and getattr(str_ref, "f", None):
            return str_ref.f
        return None

    @staticmethod
    def _parse_range_formula(formula: str) -> DataRange:
        """Parse "'Sheet1'!$B$2:$B$10" into a DataRange."""
        sheet_part, rng = formula.split("!")
        sheet_name = sheet_part.strip("'")
        clean = rng.replace("$", "")
        if ":" in clean:
            start, end = clean.split(":")
        else:
            start = end = clean
        return DataRange(sheet_name=sheet_name, start=start, end=end)

    # ---- live cell reading (replaces cache-only approach) ---------------------

    @staticmethod
    def _cells_from_range(wb: Workbook, formula: str) -> List:
        """
        Resolve a range formula like "'Sheet1'!$B$2:$B$10" against the
        actual workbook and return a flat list of cell values.
        """
        sheet_part, rng = formula.split("!")
        sheet_name = sheet_part.strip("'")
        ws = wb[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(rng.replace("$", ""))
        out = []
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                out.append(ws.cell(row=r, column=c).value)
        return out

    def _read_labels(self, wb: Workbook, formula: str) -> List[str]:
        """Read cell values from a range and return them as label strings."""
        raw = self._cells_from_range(wb, formula)
        return [str(v) if v is not None else "" for v in raw]

    def _read_values(self, wb: Workbook, formula: str) -> List[Any]:
        """Read cell values from a range and return them as-is."""
        return self._cells_from_range(wb, formula)

    # ---- series label (tx) helper ---------------------------------------------

    def _read_series_label(self, wb: Workbook, tx) -> Optional[str]:
        """
        Extract a series label from the series' tx (SeriesLabel) object.
        Tries the live workbook first, falls back to an inline literal.
        """
        if tx is None:
            return None
        str_ref = getattr(tx, "strRef", None)
        if str_ref and getattr(str_ref, "f", None):
            cells = self._cells_from_range(wb, str_ref.f)
            if cells:
                return str(cells[0]) if cells[0] is not None else ""
        # inline literal value
        if getattr(tx, "v", None) is not None:
            return str(tx.v)
        return None

    # ---- series extraction ----------------------------------------------------

    def _extract_categories(
        self, chart, wb: Workbook
    ) -> tuple[List[str], Optional[DataRange]]:
        """
        Extract category labels from the first series that has a cat reference.
        Returns (labels, DataRange) or ([], None).
        """
        for series in getattr(chart, "series", []):
            cat = getattr(series, "cat", None)
            if cat:
                ref = self._ref_formula(cat)
                if ref:
                    return self._read_labels(wb, ref), self._parse_range_formula(ref)
        return [], None

    def _extract_all_series(self, chart, wb: Workbook) -> List[ChartSeries]:
        """
        Walk every series on the chart and build a ChartSeries for each,
        resolving values from the live workbook.
        """
        result: List[ChartSeries] = []

        for series in getattr(chart, "series", []):
            name = self._read_series_label(wb, getattr(series, "tx", None))

            data_range: Optional[DataRange] = None
            values: List[Any] = []
            val = getattr(series, "val", None)
            if val:
                ref = self._ref_formula(val)
                if ref:
                    data_range = self._parse_range_formula(ref)
                    values = self._read_values(wb, ref)

            result.append(
                ChartSeries(
                    name=name,
                    data_range=data_range,
                    values=values,
                )
            )

        return result

    # ---- public API -----------------------------------------------------------

    def extract(self, sheet: Worksheet, wb: Workbook) -> List[ChartData]:
        """Extract all charts from a worksheet and return structured ChartData."""
        charts: List[ChartData] = []

        for ch in getattr(sheet, "_charts", []):
            categories, category_range = self._extract_categories(ch, wb)
            all_series = self._extract_all_series(ch, wb)

            charts.append(
                ChartData(
                    title=self._extract_title(ch.title),
                    x_axis=self._extract_axis_title(ch, "x_axis"),
                    y_axis=self._extract_axis_title(ch, "y_axis"),
                    bounding_box=self._extract_bounding_box(ch),
                    chart_type=self._extract_chart_type(ch),
                    categories=categories,
                    category_range=category_range,
                    series=all_series,
                )
            )

        return charts
