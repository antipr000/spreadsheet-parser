"""
SheetExtractor — the main per-sheet orchestrator.

Responsibilities:
  1. Read & normalise all cells in the worksheet (merged cells, data
     validations, ArrayFormula handling, actual-used-range detection).
  2. Split the used range into candidate regions via empty-row/col gaps.
  3. Build a ``RegionData`` for each candidate.
  4. Run the detection chain (Heading → KeyValue → Text → Table) on each
     region, respecting the ``DETECTION_TYPE`` setting.
  5. Extract charts separately and wrap them as ``ChartBlock`` objects.
  6. Return all blocks in reading order.
"""

from __future__ import annotations

import logging
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import Cell

from detection.base import Detector
from detection.constants import DETECTION_TYPE
from detection import (
    HeadingDetector,
    KeyValueDetector,
    TextDetector,
    TableDetector,
)
from ai.factory import get_decision_service
from ai.response_parser import parse_llm_json
from dto.blocks import Block, ChartBlock
from dto.cell_data import CellData
from dto.coordinate import BoundingBox
from dto.region import RegionData
from extractors.chart import ChartExtractor
from prompts.region_split import get_region_refinement_prompt

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _coord(col: int, row: int) -> str:
    """Return an A1-style coordinate from 1-based col/row indices."""
    return f"{get_column_letter(col)}{row}"


def _color_hex(color_obj) -> Optional[str]:
    """Best-effort extraction of an RGB hex string from an openpyxl color."""
    if color_obj is None:
        return None
    if color_obj.type == "rgb" and color_obj.rgb and str(color_obj.rgb) != "00000000":
        rgb = str(color_obj.rgb)
        return f"#{rgb[-6:]}" if len(rgb) >= 6 else None
    if color_obj.type == "theme":
        return f"theme:{color_obj.theme}"
    return None


# =====================================================================
# SheetExtractor
# =====================================================================


class SheetExtractor:
    """
    Extracts all semantic blocks from a single worksheet.

    Usage::

        extractor = SheetExtractor()
        blocks = extractor.extract(ws, wb)
    """

    # The canonical detector chain — evaluated in this order.
    # The first detector that returns a Block wins.
    _DETECTORS: List[Detector] = [
        HeadingDetector(),
        KeyValueDetector(),
        TextDetector(),
        TableDetector(),
    ]

    # ------------------------------------------------------------------
    # 1.  Read & normalise cells
    # ------------------------------------------------------------------

    def _find_actual_used_range(self, ws: Worksheet) -> Tuple[int, int, int, int]:
        """
        Return (min_row, min_col, max_row, max_col) — all 1-based.

        Uses ``ws.calculate_dimension()`` with a sanity cap, then falls
        back to scanning for non-empty cells.
        """
        dim = ws.calculate_dimension()
        if dim and dim != "A1:A1":
            try:
                parts = dim.replace("$", "").split(":")
                if len(parts) == 2:
                    tl, br = parts
                    tl_col = column_index_from_string(
                        "".join(c for c in tl if c.isalpha())
                    )
                    br_col = column_index_from_string(
                        "".join(c for c in br if c.isalpha())
                    )
                    tl_row = int("".join(c for c in tl if c.isdigit()))
                    br_row = int("".join(c for c in br if c.isdigit()))
                    if (br_row - tl_row + 1) * (br_col - tl_col + 1) <= 500_000:
                        return tl_row, tl_col, br_row, br_col
            except Exception:
                pass

        min_r = min_c = float("inf")
        max_r = max_c = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    min_r = min(min_r, cell.row)
                    max_r = max(max_r, cell.row)
                    min_c = min(min_c, cell.column)
                    max_c = max(max_c, cell.column)
        if max_r == 0:
            return 1, 1, 1, 1
        return int(min_r), int(min_c), int(max_r), int(max_c)

    def _build_merge_map(self, ws: Worksheet) -> Dict[str, str]:
        merge_map: Dict[str, str] = {}
        for merged_range in ws.merged_cells.ranges:
            tl_coord = _coord(merged_range.min_col, merged_range.min_row)
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    coord = _coord(col, row)
                    if coord != tl_coord:
                        merge_map[coord] = tl_coord
        return merge_map

    def _build_validation_map(self, ws: Worksheet) -> Dict[str, List[str]]:
        val_map: Dict[str, List[str]] = {}
        try:
            for dv in ws.data_validations.dataValidation:
                if dv.type == "list" and dv.formula1:
                    raw = dv.formula1.strip('"')
                    choices = [v.strip() for v in raw.split(",") if v.strip()]
                    for cell_range in dv.sqref.ranges:
                        for row in range(cell_range.min_row, cell_range.max_row + 1):
                            for col in range(
                                cell_range.min_col, cell_range.max_col + 1
                            ):
                                val_map[_coord(col, row)] = choices
        except Exception:
            pass
        return val_map

    def _read_cell(
        self,
        cell: Cell,
        merge_map: Dict[str, str],
        val_map: Dict[str, List[str]],
        computed_values: Optional[Dict[Tuple[str, str], Any]] = None,
        sheet_name_upper: str = "",
    ) -> CellData:
        coord = _coord(cell.column, cell.row)

        value = cell.value
        merged_with = merge_map.get(coord)
        if merged_with is not None and value is None:
            pass

        # Extract formula and resolve the display value.
        # For formula cells, look up the computed value from the pre-calculated
        # formula results (keyed by (SHEET_NAME, COORDINATE)).
        formula: Optional[str] = None
        if isinstance(value, ArrayFormula):
            formula_text = getattr(value, "text", None) or ""
            formula = f"{{{formula_text}}}"
            # Look up computed value
            cv = (computed_values or {}).get((sheet_name_upper, coord.upper()))
            if cv is not None:
                value = cv
            else:
                value = formula  # fallback: show formula as value
        elif isinstance(value, str) and value.startswith("="):
            formula = value
            # Look up computed value
            cv = (computed_values or {}).get((sheet_name_upper, coord.upper()))
            if cv is not None:
                value = cv

        str_value: Optional[str] = None
        if value is not None:
            str_value = str(value)

        font = cell.font
        fill = cell.fill

        bg_color: Optional[str] = None
        if fill and fill.fgColor:
            bg_color = _color_hex(fill.fgColor)

        font_color: Optional[str] = None
        if font and font.color:
            font_color = _color_hex(font.color)

        return CellData(
            coordinate=coord,
            value=str_value,
            formula=formula,
            background_color=bg_color,
            font_color=font_color,
            font_size=font.size if font else None,
            font_name=font.name if font else None,
            font_bold=font.bold if font else None,
            font_italic=font.italic if font else None,
            font_underline=(
                True if font and font.underline and font.underline != "none" else None
            ),
            font_strikethrough=font.strikethrough if font else None,
            font_subscript=(
                font.vertAlign == "subscript" if font and font.vertAlign else None
            ),
            font_superscript=(
                font.vertAlign == "superscript" if font and font.vertAlign else None
            ),
            merged_with=merged_with,
            data_validation=val_map.get(coord),
        )

    def _read_all_cells(
        self,
        ws: Worksheet,
        computed_values: Optional[Dict[Tuple[str, str], Any]] = None,
    ) -> Tuple[List[CellData], int, int, int, int]:
        min_row, min_col, max_row, max_col = self._find_actual_used_range(ws)
        merge_map = self._build_merge_map(ws)
        val_map = self._build_validation_map(ws)
        sheet_name_upper = (ws.title or "").upper()

        cells: List[CellData] = []
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cells.append(
                    self._read_cell(
                        cell,
                        merge_map,
                        val_map,
                        computed_values=computed_values,
                        sheet_name_upper=sheet_name_upper,
                    )
                )
        return cells, min_row, min_col, max_row, max_col

    # ------------------------------------------------------------------
    # 2.  Region splitting  (empty-gap heuristic)
    # ------------------------------------------------------------------

    @staticmethod
    def _build_grid(
        cells: List[CellData],
    ) -> Dict[Tuple[int, int], CellData]:
        grid: Dict[Tuple[int, int], CellData] = {}
        for cd in cells:
            col_str = "".join(c for c in cd.coordinate if c.isalpha())
            row_num = int("".join(c for c in cd.coordinate if c.isdigit()))
            col_num = column_index_from_string(col_str)
            grid[(row_num, col_num)] = cd
        return grid

    @staticmethod
    def _is_row_empty(
        grid: Dict[Tuple[int, int], CellData],
        row: int,
        min_col: int,
        max_col: int,
    ) -> bool:
        for col in range(min_col, max_col + 1):
            cd = grid.get((row, col))
            if cd and cd.value is not None:
                return False
        return True

    @staticmethod
    def _is_col_empty(
        grid: Dict[Tuple[int, int], CellData],
        col: int,
        min_row: int,
        max_row: int,
    ) -> bool:
        for row in range(min_row, max_row + 1):
            cd = grid.get((row, col))
            if cd and cd.value is not None:
                return False
        return True

    def _split_into_regions(
        self,
        grid: Dict[Tuple[int, int], CellData],
        min_row: int,
        min_col: int,
        max_row: int,
        max_col: int,
    ) -> List[Tuple[int, int, int, int]]:
        """Return list of (min_row, min_col, max_row, max_col) tuples."""
        # Row bands
        row_bands: List[Tuple[int, int]] = []
        in_band = False
        band_start = min_row
        for r in range(min_row, max_row + 1):
            empty = self._is_row_empty(grid, r, min_col, max_col)
            if not empty and not in_band:
                band_start = r
                in_band = True
            elif empty and in_band:
                row_bands.append((band_start, r - 1))
                in_band = False
        if in_band:
            row_bands.append((band_start, max_row))

        # Column bands
        col_bands: List[Tuple[int, int]] = []
        in_band = False
        band_start = min_col
        for c in range(min_col, max_col + 1):
            empty = self._is_col_empty(grid, c, min_row, max_row)
            if not empty and not in_band:
                band_start = c
                in_band = True
            elif empty and in_band:
                col_bands.append((band_start, c - 1))
                in_band = False
        if in_band:
            col_bands.append((band_start, max_col))

        if not row_bands or not col_bands:
            return []

        regions: List[Tuple[int, int, int, int]] = []
        for r_start, r_end in row_bands:
            for c_start, c_end in col_bands:
                has_data = any(
                    grid.get((r, c)) is not None and grid[(r, c)].value is not None
                    for r in range(r_start, r_end + 1)
                    for c in range(c_start, c_end + 1)
                )
                if has_data:
                    regions.append((r_start, c_start, r_end, c_end))
        return regions

    # ------------------------------------------------------------------
    # 2b. LLM-based region refinement
    # ------------------------------------------------------------------

    def _refine_regions_with_ai(
        self,
        all_cells: List[CellData],
        heuristic_regions: List[Tuple[int, int, int, int]],
    ) -> List[Tuple[int, int, int, int]]:
        """
        Send the full sheet cell data plus the heuristic region boundaries
        to the LLM and ask it to refine — splitting regions that contain
        multiple adjacent blocks without gaps.

        Returns a refined list of (min_row, min_col, max_row, max_col).
        Falls back to the heuristic regions on any failure.
        """
        # Build A1-notation bounding boxes for the prompt
        heuristic_boxes = [
            (_coord(c_min, r_min), _coord(c_max, r_max))
            for r_min, c_min, r_max, c_max in heuristic_regions
        ]

        prompt = get_region_refinement_prompt(all_cells, heuristic_boxes)
        ai = get_decision_service()

        try:
            raw = ai.get_decision(prompt)
        except Exception:
            logger.warning(
                "LLM region refinement call failed — using heuristic regions",
                exc_info=True,
            )
            return heuristic_regions

        parsed = parse_llm_json(raw)
        if not isinstance(parsed, list):
            logger.warning(
                "LLM region refinement did not return a JSON array — using heuristic regions"
            )
            return heuristic_regions

        refined: List[Tuple[int, int, int, int]] = []
        for item in parsed:
            try:
                tl = item.get("top_left", "")
                br = item.get("bottom_right", "")

                tl_col_str = "".join(c for c in tl if c.isalpha())
                tl_row = int("".join(c for c in tl if c.isdigit()))
                br_col_str = "".join(c for c in br if c.isalpha())
                br_row = int("".join(c for c in br if c.isdigit()))

                tl_col = column_index_from_string(tl_col_str)
                br_col = column_index_from_string(br_col_str)

                if tl_row <= br_row and tl_col <= br_col:
                    refined.append((tl_row, tl_col, br_row, br_col))
                else:
                    logger.warning("Skipping invalid refined region: %s to %s", tl, br)
            except Exception as exc:
                logger.warning("Skipping unparseable refined region %s: %s", item, exc)

        if not refined:
            logger.warning(
                "LLM region refinement returned no valid regions — using heuristic regions"
            )
            return heuristic_regions

        logger.info(
            "  Region refinement: %d heuristic → %d refined",
            len(heuristic_regions),
            len(refined),
        )
        return refined

    # ------------------------------------------------------------------
    # 3.  Build RegionData from bounds
    # ------------------------------------------------------------------

    @staticmethod
    def _make_region(
        grid: Dict[Tuple[int, int], CellData],
        r_min: int,
        c_min: int,
        r_max: int,
        c_max: int,
    ) -> RegionData:
        cells: List[CellData] = []
        sub_grid: Dict[Tuple[int, int], CellData] = {}
        for r in range(r_min, r_max + 1):
            for c in range(c_min, c_max + 1):
                cd = grid.get((r, c))
                if cd:
                    cells.append(cd)
                    sub_grid[(r, c)] = cd

        return RegionData(
            cells=cells,
            bounding_box=BoundingBox(
                top_left=_coord(c_min, r_min),
                bottom_right=_coord(c_max, r_max),
            ),
            min_row=r_min,
            min_col=c_min,
            max_row=r_max,
            max_col=c_max,
            grid=sub_grid,
        )

    # ------------------------------------------------------------------
    # 4.  Detection dispatch
    # ------------------------------------------------------------------

    def _run_detection(self, region: RegionData) -> Optional[Block]:
        """
        Run the detector chain on a region.

        Uses ``DETECTION_TYPE`` from constants to decide which method(s)
        to call:
          - ``"heuristic"``         — heuristic only
          - ``"ai"``                — AI only
          - ``"heuristic_then_ai"`` — try heuristic first, fall back to AI
        """
        for detector in self._DETECTORS:
            block: Optional[Block] = None

            if DETECTION_TYPE == "heuristic":
                block = detector.detect(region)
            elif DETECTION_TYPE == "ai":
                block = detector.detect_with_ai(region)
            elif DETECTION_TYPE == "heuristic_then_ai":
                block = detector.detect(region)
                if block is None:
                    block = detector.detect_with_ai(region)

            if block is not None:
                return block

        return None

    # ------------------------------------------------------------------
    # 5.  Chart extraction  → ChartBlock
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_chart_blocks(ws: Worksheet, wb: Workbook) -> List[Block]:
        chart_extractor = ChartExtractor()
        chart_datas = chart_extractor.extract(ws, wb)
        blocks: List[Block] = []
        for cd in chart_datas:
            blocks.append(
                ChartBlock(
                    bounding_box=cd.bounding_box,
                    chart_data=cd,
                )
            )
        return blocks

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def extract(
        self,
        ws: Worksheet,
        wb: Workbook,
        computed_values: Optional[Dict[Tuple[str, str], Any]] = None,
    ) -> List[Block]:
        """
        Extract all semantic blocks from a worksheet.

        Args:
            ws: The worksheet from the formula workbook (data_only=False).
            wb: The formula workbook (needed for chart series resolution).
            computed_values: Pre-computed formula results keyed by
                ``(SHEET_NAME_UPPER, COORDINATE_UPPER)``.

        Returns blocks sorted in reading order (top→bottom, left→right).
        """
        # Step 1: Read cells
        all_cells, min_row, min_col, max_row, max_col = self._read_all_cells(
            ws, computed_values
        )
        if not all_cells:
            return self._extract_chart_blocks(ws, wb)

        grid = self._build_grid(all_cells)

        # Step 2: Heuristic split into candidate regions (whitespace gaps)
        region_bounds = self._split_into_regions(
            grid, min_row, min_col, max_row, max_col
        )

        # Step 2b: Refine regions with LLM to split
        # adjacent blocks that have no whitespace gap between them.
        region_bounds = self._refine_regions_with_ai(all_cells, region_bounds)

        # Step 3 + 4: For each region, run detection chain
        blocks: List[Block] = []
        for r_min, c_min, r_max, c_max in region_bounds:
            region = self._make_region(grid, r_min, c_min, r_max, c_max)
            if not region.non_empty_cells:
                continue
            try:
                block = self._run_detection(region)
                if block is not None:
                    blocks.append(block)
                else:
                    logger.debug(
                        "No detector matched region %s — skipping",
                        region.bounding_box,
                    )
            except Exception:
                logger.exception(
                    "Detection failed for region %s — skipping",
                    region.bounding_box,
                )

        # Step 5: Charts
        blocks.extend(self._extract_chart_blocks(ws, wb))

        # Sort in reading order: top-to-bottom, then left-to-right
        blocks.sort(key=self._block_sort_key)
        return blocks

    @staticmethod
    def _block_sort_key(block: Block) -> Tuple[int, int]:
        """Sort key: (top_left_row, top_left_col)."""
        coord = block.bounding_box.top_left
        col_str = "".join(c for c in coord if c.isalpha())
        row_num = int("".join(c for c in coord if c.isdigit()) or "0")
        try:
            col_num = column_index_from_string(col_str) if col_str else 0
        except Exception:
            col_num = 0
        return (row_num, col_num)
