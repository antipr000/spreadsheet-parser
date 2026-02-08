"""
Extracts tables from a worksheet.

Pipeline:
  1. Read & normalise every cell in the used range (handling merged cells,
     data validations, actual-used-range detection).
  2. Heuristic detection – split the sheet into candidate table regions by
     finding fully-empty row / column gaps.
  3. LLM refinement – for each candidate region, ask the LLM to confirm or
     split further, and to identify header / body / footer sections.
  4. Post-process & validate – build the final TableData DTOs.
"""

from __future__ import annotations

import logging
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import Cell

from ai.factory import get_decision_service
from ai.response_parser import parse_llm_json
from dto.ai import TableSchemaDTO
from dto.cell_data import CellData
from dto.coordinate import BoundingBox
from dto.table_data import TableData
from prompts.bounding_box import get_bounding_box_prompt

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _coord(col: int, row: int) -> str:
    """Return an A1-style coordinate from 1-based col/row indices."""
    return f"{get_column_letter(col)}{row}"


def _color_hex(color_obj) -> Optional[str]:
    """Best-effort extraction of an RGB hex string from an openpyxl color."""
    if color_obj is None:
        return None
    if color_obj.type == "rgb" and color_obj.rgb and str(color_obj.rgb) != "00000000":
        rgb = str(color_obj.rgb)
        # openpyxl may return ARGB (8 chars); strip the alpha channel
        return f"#{rgb[-6:]}" if len(rgb) >= 6 else None
    if color_obj.type == "theme":
        return f"theme:{color_obj.theme}"
    return None


# =====================================================================
# TableExtractor
# =====================================================================


class TableExtractor:

    # ------------------------------------------------------------------
    # 1.  Read & normalise cells
    # ------------------------------------------------------------------

    def _find_actual_used_range(self, ws: Worksheet) -> Tuple[int, int, int, int]:
        """
        Walk the sheet to find the actual min/max row & col that contain
        data.  ws.calculate_dimension() can be stale / overly large, so
        we fall back to scanning ws.iter_rows when needed.

        Returns (min_row, min_col, max_row, max_col) – all 1-based.
        """
        dim = ws.calculate_dimension()  # e.g. "A1:Z100"
        if dim and dim != "A1:A1":
            # Quick sanity check: if the range is suspiciously huge (>500k
            # cells), scan instead.

            try:
                parts = dim.replace("$", "").split(":")
                if len(parts) == 2:
                    tl, br = parts
                    tl_col = column_index_from_string(
                        "".join(c for c in tl if c.isalpha())
                    )
                    br_col = column_index_from_string(
                        "".join(c for c in br if c.isalpha())
                    )
                    tl_row = int("".join(c for c in tl if c.isdigit()))
                    br_row = int("".join(c for c in br if c.isdigit()))
                    total_cells = (br_row - tl_row + 1) * (br_col - tl_col + 1)
                    if total_cells <= 500_000:
                        return tl_row, tl_col, br_row, br_col
            except Exception:
                pass

        # Fallback: scan for actual non-empty cells
        min_r = min_c = float("inf")
        max_r = max_c = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    min_r = min(min_r, cell.row)
                    max_r = max(max_r, cell.row)
                    min_c = min(min_c, cell.column)
                    max_c = max(max_c, cell.column)
        if max_r == 0:
            return 1, 1, 1, 1
        return int(min_r), int(min_c), int(max_r), int(max_c)

    def _build_merge_map(self, ws: Worksheet) -> Dict[str, str]:
        """
        Return a mapping  cell_coord → top_left_coord for every cell
        that is part of a merged range (excluding the top-left master cell
        itself).
        """
        merge_map: Dict[str, str] = {}
        for merged_range in ws.merged_cells.ranges:
            top_left = merged_range.min_row, merged_range.min_col
            tl_coord = _coord(top_left[1], top_left[0])
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    coord = _coord(col, row)
                    if coord != tl_coord:
                        merge_map[coord] = tl_coord
        return merge_map

    def _build_validation_map(self, ws: Worksheet) -> Dict[str, List[str]]:
        """
        Return a mapping  cell_coord → list of allowed values for cells
        that have an in-cell dropdown / list validation.
        """
        val_map: Dict[str, List[str]] = {}
        try:
            for dv in ws.data_validations.dataValidation:
                if dv.type == "list" and dv.formula1:
                    # formula1 is either a comma-separated literal list
                    # (e.g. '"Yes,No"') or a range reference.
                    raw = dv.formula1.strip('"')
                    choices = [v.strip() for v in raw.split(",") if v.strip()]
                    for cell_range in dv.sqref.ranges:
                        for row in range(cell_range.min_row, cell_range.max_row + 1):
                            for col in range(
                                cell_range.min_col, cell_range.max_col + 1
                            ):
                                val_map[_coord(col, row)] = choices
        except Exception:
            pass  # worksheet may have no data_validations attribute
        return val_map

    def _read_cell(
        self,
        cell: Cell,
        merge_map: Dict[str, str],
        val_map: Dict[str, List[str]],
    ) -> CellData:
        """Convert a single openpyxl Cell to our CellData DTO."""
        coord = _coord(cell.column, cell.row)

        # If cell is part of a merge, pull value from the master cell
        value = cell.value
        merged_with = merge_map.get(coord)
        if merged_with is not None and value is None:
            # The master cell has the value; openpyxl returns None for
            # non-master cells in a merge range.  We record the master
            # coordinate so downstream code knows this is a merged cell.
            pass

        # Resolve ArrayFormula objects into their formula text.
        # When data_only=False, openpyxl returns ArrayFormula instances for
        # CSE / dynamic-array cells instead of plain strings.
        formula: Optional[str] = None
        if isinstance(value, ArrayFormula):
            formula_text = getattr(value, "text", None) or ""
            formula = f"{{{formula_text}}}"  # wrap in braces to denote array formula
            value = formula  # use the formula string as the display value

        # Normal string formulas (e.g. "=SUM(A1:A10)")
        elif isinstance(value, str) and value.startswith("="):
            formula = value

        # Resolve value to string
        str_value: Optional[str] = None
        if value is not None:
            str_value = str(value)

        # Font properties
        font = cell.font
        fill = cell.fill

        bg_color: Optional[str] = None
        if fill and fill.fgColor:
            bg_color = _color_hex(fill.fgColor)

        font_color: Optional[str] = None
        if font and font.color:
            font_color = _color_hex(font.color)

        return CellData(
            coordinate=coord,
            value=str_value,
            formula=formula,
            background_color=bg_color,
            font_color=font_color,
            font_size=font.size if font else None,
            font_name=font.name if font else None,
            font_bold=font.bold if font else None,
            font_italic=font.italic if font else None,
            font_underline=(
                True if font and font.underline and font.underline != "none" else None
            ),
            font_strikethrough=font.strikethrough if font else None,
            font_subscript=(
                font.vertAlign == "subscript" if font and font.vertAlign else None
            ),
            font_superscript=(
                font.vertAlign == "superscript" if font and font.vertAlign else None
            ),
            merged_with=merged_with,
            data_validation=val_map.get(coord),
        )

    def _read_all_cells(
        self, ws: Worksheet
    ) -> Tuple[List[CellData], int, int, int, int]:
        """
        Read every cell in the actual used range and return
        (cells, min_row, min_col, max_row, max_col).
        """
        min_row, min_col, max_row, max_col = self._find_actual_used_range(ws)
        merge_map = self._build_merge_map(ws)
        val_map = self._build_validation_map(ws)

        cells: List[CellData] = []
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cells.append(self._read_cell(cell, merge_map, val_map))
        return cells, min_row, min_col, max_row, max_col

    # ------------------------------------------------------------------
    # 2.  Heuristic table detection  (empty-gap splitting)
    # ------------------------------------------------------------------

    @staticmethod
    def _build_occupancy_grid(
        cells: List[CellData],
        min_row: int,
        min_col: int,
        max_row: int,
        max_col: int,
    ) -> Dict[Tuple[int, int], CellData]:
        """Index cells by (row, col) for O(1) lookup."""
        grid: Dict[Tuple[int, int], CellData] = {}
        for cd in cells:
            col_str = "".join(c for c in cd.coordinate if c.isalpha())
            row_num = int("".join(c for c in cd.coordinate if c.isdigit()))
            col_num = column_index_from_string(col_str)
            grid[(row_num, col_num)] = cd
        return grid

    @staticmethod
    def _is_row_empty(
        grid: Dict[Tuple[int, int], CellData],
        row: int,
        min_col: int,
        max_col: int,
    ) -> bool:
        for col in range(min_col, max_col + 1):
            cd = grid.get((row, col))
            if cd and cd.value is not None:
                return False
        return True

    @staticmethod
    def _is_col_empty(
        grid: Dict[Tuple[int, int], CellData],
        col: int,
        min_row: int,
        max_row: int,
    ) -> bool:
        for row in range(min_row, max_row + 1):
            cd = grid.get((row, col))
            if cd and cd.value is not None:
                return False
        return True

    def _split_into_candidate_regions(
        self,
        cells: List[CellData],
        min_row: int,
        min_col: int,
        max_row: int,
        max_col: int,
    ) -> List[Tuple[int, int, int, int]]:
        """
        Split the used range into rectangular candidate regions by detecting
        fully-empty rows and columns as separators.

        Returns a list of (min_row, min_col, max_row, max_col) tuples.
        """
        grid = self._build_occupancy_grid(cells, min_row, min_col, max_row, max_col)

        # Find non-empty row bands
        row_bands: List[Tuple[int, int]] = []
        in_band = False
        band_start = min_row
        for r in range(min_row, max_row + 1):
            empty = self._is_row_empty(grid, r, min_col, max_col)
            if not empty and not in_band:
                band_start = r
                in_band = True
            elif empty and in_band:
                row_bands.append((band_start, r - 1))
                in_band = False
        if in_band:
            row_bands.append((band_start, max_row))

        # Find non-empty column bands
        col_bands: List[Tuple[int, int]] = []
        in_band = False
        band_start = min_col
        for c in range(min_col, max_col + 1):
            empty = self._is_col_empty(grid, c, min_row, max_row)
            if not empty and not in_band:
                band_start = c
                in_band = True
            elif empty and in_band:
                col_bands.append((band_start, c - 1))
                in_band = False
        if in_band:
            col_bands.append((band_start, max_col))

        if not row_bands or not col_bands:
            return []

        # Each combination of row-band × col-band is a candidate region,
        # but only if it actually contains data.
        regions: List[Tuple[int, int, int, int]] = []
        for r_start, r_end in row_bands:
            for c_start, c_end in col_bands:
                has_data = False
                for r in range(r_start, r_end + 1):
                    for c in range(c_start, c_end + 1):
                        cd = grid.get((r, c))
                        if cd and cd.value is not None:
                            has_data = True
                            break
                    if has_data:
                        break
                if has_data:
                    regions.append((r_start, c_start, r_end, c_end))
        return regions

    def _cells_in_region(
        self,
        grid: Dict[Tuple[int, int], CellData],
        r_min: int,
        c_min: int,
        r_max: int,
        c_max: int,
    ) -> List[CellData]:
        """Return all CellData objects within a rectangular region."""
        out: List[CellData] = []
        for r in range(r_min, r_max + 1):
            for c in range(c_min, c_max + 1):
                cd = grid.get((r, c))
                if cd:
                    out.append(cd)
        return out

    # ------------------------------------------------------------------
    # 3.  LLM refinement
    # ------------------------------------------------------------------

    def _call_llm(self, region_cells: List[CellData]) -> List[TableSchemaDTO]:
        """
        Send the cell data for a candidate region to the LLM and parse the
        response into a list of ``TableSchemaDTO`` objects.
        """
        prompt = get_bounding_box_prompt(region_cells)
        ai = get_decision_service()
        raw_response = ai.get_decision(prompt)

        parsed = parse_llm_json(raw_response)
        if not isinstance(parsed, list):
            logger.warning("LLM response was not a JSON array: %s", raw_response[:200])
            return []

        schemas: List[TableSchemaDTO] = []
        for item in parsed:
            try:
                schemas.append(TableSchemaDTO.model_validate(item))
            except Exception as exc:
                logger.warning(
                    "Skipping invalid table schema from LLM: %s — %s", exc, item
                )
        return schemas

    # ------------------------------------------------------------------
    # 4.  Validation & post-processing
    # ------------------------------------------------------------------

    @staticmethod
    def _validate_schema(
        schema: TableSchemaDTO,
        min_row: int,
        min_col: int,
        max_row: int,
        max_col: int,
    ) -> bool:
        """
        Sanity-check a ``TableSchemaDTO`` against the known sheet bounds.
        Returns True if the schema looks plausible.
        """
        try:
            tl_col = column_index_from_string(
                "".join(c for c in schema.top_left if c.isalpha())
            )
            tl_row = int("".join(c for c in schema.top_left if c.isdigit()))
            br_col = column_index_from_string(
                "".join(c for c in schema.bottom_right if c.isalpha())
            )
            br_row = int("".join(c for c in schema.bottom_right if c.isdigit()))
        except Exception:
            return False

        # Bounding box should be within sheet bounds (with some tolerance)
        if tl_row > br_row or tl_col > br_col:
            return False
        if br_row < min_row or tl_row > max_row:
            return False
        if br_col < min_col or tl_col > max_col:
            return False

        # header, body, footer rows should not overlap
        all_rows = (
            set(schema.header_rows) | set(schema.body_rows) | set(schema.footer_rows)
        )
        if len(all_rows) < len(schema.header_rows) + len(schema.body_rows) + len(
            schema.footer_rows
        ):
            return False

        return True

    def _build_table_data(
        self,
        schema: TableSchemaDTO,
        grid: Dict[Tuple[int, int], CellData],
    ) -> TableData:
        """
        Given a validated schema and the cell grid, assemble a ``TableData``
        with heading / data / footer cell lists.
        """
        header_col_indices = [
            column_index_from_string(c) for c in schema.header_columns
        ]
        body_col_indices = [column_index_from_string(c) for c in schema.body_columns]
        footer_col_indices = [
            column_index_from_string(c) for c in schema.footer_columns
        ]

        heading: List[CellData] = []
        for r in schema.header_rows:
            for c in header_col_indices:
                cd = grid.get((r, c))
                if cd:
                    heading.append(cd)

        data: List[CellData] = []
        for r in schema.body_rows:
            for c in body_col_indices:
                cd = grid.get((r, c))
                if cd:
                    data.append(cd)

        footer: List[CellData] = []
        for r in schema.footer_rows:
            for c in footer_col_indices:
                cd = grid.get((r, c))
                if cd:
                    footer.append(cd)

        return TableData(
            bounding_box=BoundingBox(
                top_left=schema.top_left,
                bottom_right=schema.bottom_right,
            ),
            heading=heading,
            data=data,
            footer=footer,
        )

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def extract(self, ws: Worksheet) -> List[TableData]:
        """
        Extract all tables from a worksheet.

        Returns a list of ``TableData`` objects, each containing the bounding
        box and categorised cell data (heading / data / footer).
        """
        # Step 1: Read & normalise
        all_cells, min_row, min_col, max_row, max_col = self._read_all_cells(ws)
        if not all_cells:
            return []

        grid = self._build_occupancy_grid(all_cells, min_row, min_col, max_row, max_col)

        # Step 2: Heuristic split into candidate regions
        regions = self._split_into_candidate_regions(
            all_cells, min_row, min_col, max_row, max_col
        )
        if not regions:
            return []

        # Step 3 & 4: For each candidate region, ask the LLM and build output
        tables: List[TableData] = []
        for r_min, c_min, r_max, c_max in regions:
            region_cells = self._cells_in_region(grid, r_min, c_min, r_max, c_max)
            if not region_cells:
                continue

            schemas = self._call_llm(region_cells)

            for schema in schemas:
                if not self._validate_schema(
                    schema, min_row, min_col, max_row, max_col
                ):
                    logger.warning("Discarding invalid LLM schema: %s", schema)
                    continue
                tables.append(self._build_table_data(schema, grid))

        return tables
