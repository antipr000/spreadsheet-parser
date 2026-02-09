"""
Smart cell-data summariser for the Planner agent.

Compresses a sheet's cell data into a structured text summary that fits
within ~4K-8K tokens, even for sheets with 13K+ cells.  The summary
preserves enough structural signal (headers, group labels, merged
ranges, column inventory, sampled body rows) for the LLM to identify
all blocks and their types.
"""

from __future__ import annotations

import logging
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from dto.cell_data import CellData
from prompts.bounding_box import get_cell_data_prompt
from agentic_flow.cell_reader import parse_coord

logger = logging.getLogger(__name__)

# ------------------------------------------------------------------
# Configuration
# ------------------------------------------------------------------

# Maximum rows to include in full detail for the header section
_MAX_HEADER_ROWS = 8
# Number of footer/last rows to include
_MAX_FOOTER_ROWS = 3
# Maximum sampled body rows
_MAX_SAMPLE_ROWS = 10
# For wide sheets, sample every Nth column in body rows
_MAX_COLS_FULL_DETAIL = 20
# If the total prompt would exceed this many chars, aggressively compress
_TARGET_CHARS = 24_000  # ~6K tokens


# ------------------------------------------------------------------
# Row analysis helpers
# ------------------------------------------------------------------

def _row_cells(
    grid: Dict[Tuple[int, int], CellData],
    row: int,
    min_col: int,
    max_col: int,
) -> List[CellData]:
    """Return all cells (including empty) for a row within column range."""
    return [
        grid[(row, c)]
        for c in range(min_col, max_col + 1)
        if (row, c) in grid
    ]


def _non_empty_in_row(
    grid: Dict[Tuple[int, int], CellData],
    row: int,
    min_col: int,
    max_col: int,
) -> List[CellData]:
    """Return non-empty cells for a row."""
    return [
        cd for cd in _row_cells(grid, row, min_col, max_col)
        if cd.value is not None
    ]


def _is_structural_row(
    grid: Dict[Tuple[int, int], CellData],
    row: int,
    min_col: int,
    max_col: int,
    total_cols: int,
) -> Tuple[bool, str]:
    """
    Detect if a row is a "structural break" — a row that signals a
    boundary between blocks or a group label.

    Returns (is_structural, reason_string).
    """
    non_empty = _non_empty_in_row(grid, row, min_col, max_col)
    if not non_empty:
        return False, ""

    n_filled = len(non_empty)
    bold_count = sum(1 for c in non_empty if c.font_bold)
    all_bold = bold_count == n_filled and n_filled > 0

    # Single-value bold row — almost certainly a group label
    if all_bold and n_filled <= 2:
        return True, f"ONLY {n_filled} cell(s) in row, bold -- likely GROUP LABEL"

    # Bold row with much fewer columns than typical (subtotal/section break)
    if all_bold and n_filled <= max(3, total_cols * 0.15):
        return True, f"bold, {n_filled} cols filled (sparse) -- likely SECTION BREAK"

    # Bold row with many columns — subtotal row
    if all_bold and n_filled >= total_cols * 0.4:
        return True, f"bold, {n_filled} cols filled -- likely SUBTOTAL/TOTAL"

    return False, ""


# ------------------------------------------------------------------
# Merged cell analysis
# ------------------------------------------------------------------

def _merged_ranges_summary(ws: Worksheet) -> List[str]:
    """
    Summarise merged cell ranges, flagging those that span many rows
    (group-indicator columns).
    """
    lines: List[str] = []
    for mr in ws.merged_cells.ranges:
        span_rows = mr.max_row - mr.min_row + 1
        span_cols = mr.max_col - mr.min_col + 1
        tl = f"{get_column_letter(mr.min_col)}{mr.min_row}"
        br = f"{get_column_letter(mr.max_col)}{mr.max_row}"

        # Read top-left value
        val = ws.cell(row=mr.min_row, column=mr.min_col).value
        val_str = repr(val)[:50] if val is not None else "None"

        tag = ""
        if span_rows >= 3 and span_cols == 1:
            tag = "  <-- GROUP COLUMN"
        elif span_cols >= 3 and span_rows == 1:
            tag = "  <-- MULTI-COL HEADER"

        lines.append(
            f"{tl}:{br} (merged {span_rows}r x {span_cols}c, val={val_str}){tag}"
        )
    return lines


# ------------------------------------------------------------------
# Column inventory
# ------------------------------------------------------------------

def _column_inventory(
    grid: Dict[Tuple[int, int], CellData],
    header_rows: List[int],
    min_col: int,
    max_col: int,
) -> List[str]:
    """
    List each column with its header value(s) for orientation.
    """
    lines: List[str] = []
    for c in range(min_col, max_col + 1):
        col_letter = get_column_letter(c)
        header_vals = []
        for r in header_rows:
            cd = grid.get((r, c))
            if cd and cd.value is not None:
                header_vals.append(cd.value)
        if header_vals:
            joined = " / ".join(str(v)[:30] for v in header_vals)
            lines.append(f"  {col_letter}: {joined}")
    return lines


# ------------------------------------------------------------------
# Format a row for the prompt
# ------------------------------------------------------------------

def _compact_cell(cd: CellData) -> str:
    """Ultra-compact cell representation: [A1] val bold bg."""
    parts = [f"[{cd.coordinate}]"]
    if cd.value is not None:
        val_str = str(cd.value)[:30]
        parts.append(f'"{val_str}"')
    if cd.font_bold:
        parts.append("bold")
    if cd.merged_with:
        parts.append(f"merged→{cd.merged_with}")
    if cd.formula:
        parts.append("formula")
    return " ".join(parts)


def _format_row(
    grid: Dict[Tuple[int, int], CellData],
    row: int,
    min_col: int,
    max_col: int,
    *,
    col_sample_step: int = 1,
    max_cells: int = 20,
    annotation: str = "",
) -> str:
    """
    Format one row of cell data as a compact single line.
    """
    parts: List[str] = []
    count = 0
    for c in range(min_col, max_col + 1, col_sample_step):
        cd = grid.get((row, c))
        if cd and cd.value is not None:
            parts.append(_compact_cell(cd))
            count += 1
            if count >= max_cells:
                parts.append("...")
                break

    n_filled = len(_non_empty_in_row(grid, row, min_col, max_col))
    suffix = f"  ({n_filled} cols)"
    if annotation:
        suffix += f"  [{annotation}]"

    return f"Row {row}: {' | '.join(parts)}{suffix}" if parts else ""


# ------------------------------------------------------------------
# Public API
# ------------------------------------------------------------------

def summarise_sheet(
    grid: Dict[Tuple[int, int], CellData],
    ws: Worksheet,
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
) -> str:
    """
    Build a structural summary of the sheet for the Planner prompt.
    """
    total_rows = max_row - min_row + 1
    total_cols = max_col - min_col + 1
    non_empty_total = sum(
        1 for (r, c), cd in grid.items()
        if min_row <= r <= max_row and min_col <= c <= max_col
        and cd.value is not None
    )

    # Column sampling step for wide sheets
    col_step = max(1, total_cols // _MAX_COLS_FULL_DETAIL)

    sections: List[str] = []

    # --- Preamble ---
    tl = f"{get_column_letter(min_col)}{min_row}"
    br = f"{get_column_letter(max_col)}{max_row}"
    sections.append(
        f"Sheet: {ws.title} ({total_rows} rows x {total_cols} cols, {tl}:{br})\n"
        f"{non_empty_total} non-empty cells\n"
        f"Charts: {len(getattr(ws, '_charts', []))}"
    )

    # --- Header rows (first N rows, sampled for wide sheets) ---
    header_limit = min(_MAX_HEADER_ROWS, total_rows)
    header_rows_range = list(range(min_row, min_row + header_limit))
    header_lines: List[str] = []
    for r in header_rows_range:
        line = _format_row(
            grid, r, min_col, max_col,
            col_sample_step=col_step,
            max_cells=25,
        )
        if line:
            header_lines.append(line)
    if header_lines:
        sections.append(
            f"=== HEADER / TOP ROWS (rows {min_row}-{min_row + header_limit - 1}) ===\n"
            + "\n".join(header_lines)
        )

    # --- Structural rows ---
    structural_rows: List[Tuple[int, str]] = []
    # Median filled columns for "typical body row" estimation
    filled_counts = []
    for r in range(min_row, max_row + 1):
        ne = _non_empty_in_row(grid, r, min_col, max_col)
        if ne:
            filled_counts.append(len(ne))
    median_cols = sorted(filled_counts)[len(filled_counts) // 2] if filled_counts else total_cols

    for r in range(min_row + header_limit, max_row + 1):
        is_struct, reason = _is_structural_row(
            grid, r, min_col, max_col, median_cols
        )
        if is_struct:
            structural_rows.append((r, reason))

    if structural_rows:
        struct_lines: List[str] = []
        # Cap the number shown
        shown = structural_rows[:40]
        for r, reason in shown:
            line = _format_row(
                grid, r, min_col, max_col,
                col_sample_step=col_step,
                annotation=reason,
            )
            if line:
                struct_lines.append(line)
        if len(structural_rows) > 40:
            struct_lines.append(f"  ... and {len(structural_rows) - 40} more structural rows")
        sections.append(
            "=== STRUCTURAL ROWS (bold / group labels / subtotals) ===\n"
            + "\n".join(struct_lines)
        )

    # --- Merged cell ranges ---
    merge_lines = _merged_ranges_summary(ws)
    if merge_lines:
        shown = merge_lines[:30]
        extra = ""
        if len(merge_lines) > 30:
            extra = f"\n  ... and {len(merge_lines) - 30} more merged ranges"
        sections.append(
            "=== MERGED CELL RANGES ===\n"
            + "\n".join(shown) + extra
        )

    # --- Column inventory ---
    inv = _column_inventory(grid, header_rows_range, min_col, max_col)
    if inv:
        shown = inv[:50]
        extra = ""
        if len(inv) > 50:
            extra = f"\n  ... and {len(inv) - 50} more columns"
        sections.append(
            "=== COLUMN INVENTORY (header values) ===\n"
            + "\n".join(shown) + extra
        )

    # --- Sampled body rows ---
    # Pick evenly spaced rows from the body (excluding header & structural rows)
    structural_set: Set[int] = {r for r, _ in structural_rows}
    body_rows = [
        r for r in range(min_row + header_limit, max_row - _MAX_FOOTER_ROWS + 1)
        if r not in structural_set
        and _non_empty_in_row(grid, r, min_col, max_col)
    ]
    if body_rows:
        step = max(1, len(body_rows) // _MAX_SAMPLE_ROWS)
        sampled = body_rows[::step][:_MAX_SAMPLE_ROWS]
        sample_lines: List[str] = []
        for r in sampled:
            line = _format_row(
                grid, r, min_col, max_col,
                col_sample_step=col_step,
                max_cells=15,
            )
            if line:
                sample_lines.append(line)
        if sample_lines:
            sections.append(
                "=== SAMPLE BODY ROWS ===\n"
                + "\n".join(sample_lines)
            )

    # --- Footer / last rows ---
    footer_start = max(min_row + header_limit, max_row - _MAX_FOOTER_ROWS + 1)
    footer_lines: List[str] = []
    for r in range(footer_start, max_row + 1):
        line = _format_row(grid, r, min_col, max_col, max_cells=30)
        if line:
            footer_lines.append(line)
    if footer_lines:
        sections.append(
            f"=== FOOTER / LAST ROWS (rows {footer_start}-{max_row}) ===\n"
            + "\n".join(footer_lines)
        )

    return "\n\n".join(sections)
