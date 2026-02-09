"""
Table Extractor — two-pass approach for handling arbitrarily large tables.

Pass 1 (LLM): Send a compact structural snapshot (headers, sample rows,
              bold rows, merged ranges) and receive the table's structure
              (header/footer rows, row groups, column groups).

Pass 2 (programmatic): Read cells from the grid guided by the structure
                       map.  No LLM call needed.
"""

from __future__ import annotations

import logging
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from ai.factory import get_decision_service
from ai.response_parser import parse_llm_json
from dto.blocks import Block, TableBlock, RowGroup
from dto.cell_data import CellData
from dto.coordinate import BoundingBox
from prompts.bounding_box import get_cell_data_prompt

from agentic_flow.cell_reader import parse_coord, coord, slice_grid
from agentic_flow.dto.plan import PlannedBlock
from agentic_flow.extractors.base import BaseExtractor
from agentic_flow.prompts.table import get_table_structure_prompt

logger = logging.getLogger(__name__)

# If the table has fewer non-empty cells than this, skip the LLM
# structure-detection pass and use simple heuristics instead.
_SMALL_TABLE_THRESHOLD = 200

# Max cells to include in the structure-detection prompt.
_MAX_PROMPT_CELLS = 300

# Number of sample body rows to send to the LLM.
_SAMPLE_BODY_ROWS = 5
_LAST_ROWS = 3


class TableExtractor(BaseExtractor):

    def extract(
        self,
        planned: PlannedBlock,
        grid: Dict[Tuple[int, int], CellData],
        merge_map: Dict[str, str],
        ws: Worksheet,
        wb: Workbook,
        *,
        computed_values: Optional[Dict[Tuple[str, str], Any]] = None,
        screenshot_bytes: Optional[bytes] = None,
    ) -> List[Block]:
        bbox = planned.bounding_box
        r_min, c_min = parse_coord(bbox.top_left)
        r_max, c_max = parse_coord(bbox.bottom_right)

        sub = slice_grid(grid, r_min, c_min, r_max, c_max)
        non_empty = [cd for cd in sub.values() if cd.value is not None]

        if not non_empty:
            return []

        total_rows = r_max - r_min + 1
        total_cols = c_max - c_min + 1

        # ----------------------------------------------------------
        # Decide: small table → heuristic, large table → LLM Pass 1
        # ----------------------------------------------------------
        if len(non_empty) <= _SMALL_TABLE_THRESHOLD:
            structure = self._heuristic_structure(
                sub, r_min, c_min, r_max, c_max, planned
            )
        else:
            structure = self._llm_structure(
                sub, ws, r_min, c_min, r_max, c_max, total_rows, total_cols, planned
            )

        # ----------------------------------------------------------
        # Pass 2: Programmatic extraction using the structure map
        # ----------------------------------------------------------
        return self._build_table_block(
            sub, r_min, c_min, r_max, c_max, structure, bbox
        )

    # ==============================================================
    # Pass 1 Option A: Simple heuristic for small tables
    # ==============================================================

    def _heuristic_structure(
        self,
        sub: Dict[Tuple[int, int], CellData],
        r_min: int,
        c_min: int,
        r_max: int,
        c_max: int,
        planned: PlannedBlock,
    ) -> dict:
        """
        Detect table structure using formatting heuristics (no LLM).
        """
        hints = planned.table_hints
        header_count = hints.header_row_count if hints else 1

        # Detect header rows by bold ratio
        header_rows: List[int] = []
        for r in range(r_min, min(r_min + header_count + 2, r_max + 1)):
            row_cells = [
                sub[(r, c)] for c in range(c_min, c_max + 1)
                if (r, c) in sub and sub[(r, c)].value is not None
            ]
            if not row_cells:
                continue
            bold_ratio = sum(1 for c in row_cells if c.font_bold) / len(row_cells)
            if bold_ratio >= 0.5 and not header_rows:
                header_rows.append(r)
            elif bold_ratio >= 0.5 and header_rows and r == header_rows[-1] + 1:
                header_rows.append(r)
            else:
                break

        if not header_rows:
            header_rows = [r_min]

        # Detect row groups
        row_groups: List[dict] = []
        if hints and hints.has_row_groups:
            group_col = column_index_from_string(hints.row_group_label_column) if hints.row_group_label_column else c_min
            body_start = header_rows[-1] + 1
            current_group: Optional[dict] = None
            for r in range(body_start, r_max + 1):
                cd = sub.get((r, group_col))
                if cd and cd.value is not None and cd.font_bold:
                    # Check if this is a single-value row (group label)
                    filled = sum(
                        1 for c in range(c_min, c_max + 1)
                        if (r, c) in sub and sub[(r, c)].value is not None
                    )
                    if filled <= 3:
                        if current_group:
                            current_group["end_row"] = r - 1
                            row_groups.append(current_group)
                        current_group = {
                            "label_row": r,
                            "label": cd.value,
                            "start_row": r + 1,
                            "end_row": r_max,
                        }
            if current_group:
                current_group["end_row"] = r_max
                row_groups.append(current_group)

        return {
            "header_rows": header_rows,
            "header_structure": "single",
            "column_groups": [],
            "footer_rows": [],
            "row_group_label_column": (hints.row_group_label_column if hints else None),
            "row_groups": row_groups,
            "merged_group_columns": (hints.merged_group_columns if hints else []),
            "merged_groups": [],
        }

    # ==============================================================
    # Pass 1 Option B: LLM structure detection for large tables
    # ==============================================================

    def _llm_structure(
        self,
        sub: Dict[Tuple[int, int], CellData],
        ws: Worksheet,
        r_min: int,
        c_min: int,
        r_max: int,
        c_max: int,
        total_rows: int,
        total_cols: int,
        planned: PlannedBlock,
    ) -> dict:
        """
        Send a compact snapshot to the LLM and return the structure map.
        """
        hints = planned.table_hints
        header_count = hints.header_row_count if hints else 1

        # --- Collect cells for the prompt ---

        # Header rows (full)
        header_cells: List[CellData] = []
        header_row_nums = list(range(r_min, min(r_min + header_count, r_max + 1)))
        for r in header_row_nums:
            for c in range(c_min, c_max + 1):
                cd = sub.get((r, c))
                if cd and cd.value is not None:
                    header_cells.append(cd)

        body_start = header_row_nums[-1] + 1 if header_row_nums else r_min + 1

        # Sample body rows (first N)
        sample_cells: List[CellData] = []
        sample_count = 0
        for r in range(body_start, r_max + 1):
            row_has_data = any(
                (r, c) in sub and sub[(r, c)].value is not None
                for c in range(c_min, c_max + 1)
            )
            if row_has_data:
                for c in range(c_min, c_max + 1):
                    cd = sub.get((r, c))
                    if cd and cd.value is not None:
                        sample_cells.append(cd)
                sample_count += 1
                if sample_count >= _SAMPLE_BODY_ROWS:
                    break

        # Structural/bold rows
        structural_cells: List[CellData] = []
        for r in range(body_start, r_max + 1):
            row_cells = [
                sub[(r, c)] for c in range(c_min, c_max + 1)
                if (r, c) in sub and sub[(r, c)].value is not None
            ]
            if not row_cells:
                continue
            bold_count = sum(1 for c in row_cells if c.font_bold)
            if bold_count == len(row_cells) and bold_count > 0:
                for cd in row_cells:
                    structural_cells.append(cd)

        # Last rows
        last_cells: List[CellData] = []
        for r in range(max(body_start, r_max - _LAST_ROWS + 1), r_max + 1):
            for c in range(c_min, c_max + 1):
                cd = sub.get((r, c))
                if cd and cd.value is not None:
                    last_cells.append(cd)

        # Merged ranges within this table
        merged_lines: List[str] = []
        for mr in ws.merged_cells.ranges:
            # Check overlap with our bounding box
            if mr.max_row < r_min or mr.min_row > r_max:
                continue
            if mr.max_col < c_min or mr.min_col > c_max:
                continue
            tl = f"{get_column_letter(mr.min_col)}{mr.min_row}"
            br = f"{get_column_letter(mr.max_col)}{mr.max_row}"
            val = ws.cell(row=mr.min_row, column=mr.min_col).value
            val_str = repr(val)[:40] if val is not None else "None"
            span_r = mr.max_row - mr.min_row + 1
            span_c = mr.max_col - mr.min_col + 1
            merged_lines.append(f"{tl}:{br} ({span_r}r x {span_c}c, val={val_str})")

        merged_text = "\n".join(merged_lines) if merged_lines else ""

        tl_str = f"{get_column_letter(c_min)}{r_min}"
        br_str = f"{get_column_letter(c_max)}{r_max}"

        prompt = get_table_structure_prompt(
            header_cells=header_cells,
            sample_body_cells=sample_cells,
            structural_row_cells=structural_cells,
            last_rows_cells=last_cells,
            merged_ranges_text=merged_text,
            total_rows=total_rows,
            total_cols=total_cols,
            top_left=tl_str,
            bottom_right=br_str,
        )

        try:
            ai = get_decision_service()
            raw = ai.get_decision(prompt)
            parsed = parse_llm_json(raw)
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            logger.warning(
                "  [TableExtractor] LLM structure detection failed — using heuristic",
                exc_info=True,
            )

        # Fallback
        return self._heuristic_structure(sub, r_min, c_min, r_max, c_max, planned)

    # ==============================================================
    # Pass 2: Build TableBlock from structure
    # ==============================================================

    def _build_table_block(
        self,
        sub: Dict[Tuple[int, int], CellData],
        r_min: int,
        c_min: int,
        r_max: int,
        c_max: int,
        structure: dict,
        bbox: BoundingBox,
    ) -> List[Block]:
        """
        Programmatically read cells from the grid guided by the structure.
        """
        header_rows: List[int] = structure.get("header_rows", [r_min])
        footer_rows: List[int] = structure.get("footer_rows", [])

        header_set = set(header_rows)
        footer_set = set(footer_rows)

        # Collect cells
        heading_cells: List[CellData] = []
        data_cells: List[CellData] = []
        footer_cells: List[CellData] = []
        all_cells: List[CellData] = []

        for r in range(r_min, r_max + 1):
            for c in range(c_min, c_max + 1):
                cd = sub.get((r, c))
                if cd is None:
                    continue
                all_cells.append(cd)
                if r in header_set:
                    heading_cells.append(cd)
                elif r in footer_set:
                    footer_cells.append(cd)
                else:
                    data_cells.append(cd)

        # Build row groups
        row_groups: List[RowGroup] = self._build_row_groups(
            sub, r_min, c_min, r_max, c_max, structure
        )

        block = TableBlock(
            bounding_box=bbox,
            heading=heading_cells,
            data=data_cells,
            footer=footer_cells,
            cells=all_cells,
            row_groups=row_groups,
        )

        return [block]

    # ==============================================================
    # Row group construction
    # ==============================================================

    def _build_row_groups(
        self,
        sub: Dict[Tuple[int, int], CellData],
        r_min: int,
        c_min: int,
        r_max: int,
        c_max: int,
        structure: dict,
    ) -> List[RowGroup]:
        """
        Build a list of RowGroup objects from the structure map.
        """
        raw_groups = structure.get("row_groups", [])
        if not raw_groups:
            return []

        group_col_str = structure.get("row_group_label_column")
        group_col = column_index_from_string(group_col_str) if group_col_str else c_min

        groups: List[RowGroup] = []
        for rg in raw_groups:
            label_row = rg.get("label_row")
            label = rg.get("label", "")
            start_row = rg.get("start_row", label_row + 1 if label_row else r_min)
            end_row = rg.get("end_row", r_max)

            # Label cell
            label_cell = sub.get((label_row, group_col)) if label_row else None
            if label_cell is None:
                # Create a synthetic label cell
                label_cell = CellData(
                    coordinate=coord(group_col, label_row) if label_row else "A1",
                    value=str(label),
                )

            # Data rows
            data_row_cells: List[CellData] = []
            for r in range(start_row, min(end_row, r_max) + 1):
                for c in range(c_min, c_max + 1):
                    cd = sub.get((r, c))
                    if cd and cd.value is not None:
                        data_row_cells.append(cd)

            groups.append(
                RowGroup(
                    label=str(label),
                    label_cell=label_cell,
                    data_rows=data_row_cells,
                    children=[],
                )
            )

        # Nest merged-group columns into row groups
        merged_groups = structure.get("merged_groups", [])
        if merged_groups and groups:
            self._nest_merged_groups(groups, merged_groups, sub, c_min, c_max)

        return groups

    @staticmethod
    def _nest_merged_groups(
        parent_groups: List[RowGroup],
        merged_groups: List[dict],
        sub: Dict[Tuple[int, int], CellData],
        c_min: int,
        c_max: int,
    ) -> None:
        """
        For each merged-group column range, find the parent RowGroup that
        contains it and attach as a child.
        """
        for mg in merged_groups:
            mg_start = mg.get("start_row")
            mg_end = mg.get("end_row")
            mg_label = mg.get("label", "")
            mg_col_str = mg.get("column", "")

            if not mg_start or not mg_end:
                continue

            try:
                mg_col = column_index_from_string(mg_col_str)
            except Exception:
                continue

            label_cell = sub.get((mg_start, mg_col))
            if label_cell is None:
                label_cell = CellData(
                    coordinate=coord(mg_col, mg_start),
                    value=str(mg_label),
                )

            child_cells: List[CellData] = []
            for r in range(mg_start, mg_end + 1):
                for c in range(c_min, c_max + 1):
                    cd = sub.get((r, c))
                    if cd and cd.value is not None:
                        child_cells.append(cd)

            child_group = RowGroup(
                label=str(mg_label),
                label_cell=label_cell,
                data_rows=child_cells,
                children=[],
            )

            # Find the parent group that contains this range
            best_parent = None
            for pg in parent_groups:
                # Check if the merged group's label_row falls within this parent
                pg_label_row = parse_coord(pg.label_cell.coordinate)[0]
                pg_data_rows_coords = set()
                for dc in pg.data_rows:
                    pg_data_rows_coords.add(parse_coord(dc.coordinate)[0])
                all_pg_rows = pg_data_rows_coords | {pg_label_row}
                if mg_start in all_pg_rows or (
                    pg_label_row <= mg_start
                    and (not all_pg_rows or max(all_pg_rows) >= mg_end)
                ):
                    best_parent = pg
                    break

            if best_parent is not None:
                best_parent.children.append(child_group)
