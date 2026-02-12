"""
Cell reading utilities — refactored from extractors/sheet.py so they
can be shared by the agentic pipeline without instantiating
SheetExtractor.
"""

from __future__ import annotations

import logging
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import Cell

from dto.cell_data import CellData

logger = logging.getLogger(__name__)


# ------------------------------------------------------------------
# Coordinate helpers
# ------------------------------------------------------------------

def coord(col: int, row: int) -> str:
    """Return an A1-style coordinate from 1-based col/row indices."""
    return f"{get_column_letter(col)}{row}"


def parse_coord(coordinate: str) -> Tuple[int, int]:
    """Parse 'AB12' → (row=12, col=28).  Both 1-based."""
    col_str = "".join(c for c in coordinate if c.isalpha())
    row_num = int("".join(c for c in coordinate if c.isdigit()) or "0")
    col_num = column_index_from_string(col_str) if col_str else 0
    return row_num, col_num


# ------------------------------------------------------------------
# Color / fill helpers
# ------------------------------------------------------------------

def _color_hex(color_obj) -> Optional[str]:
    if color_obj is None:
        return None
    try:
        if (
            color_obj.type == "rgb"
            and color_obj.rgb
            and str(color_obj.rgb) != "00000000"
        ):
            rgb = str(color_obj.rgb)
            return f"#{rgb[-6:]}" if len(rgb) >= 6 else None
        if color_obj.type == "theme":
            return f"theme:{color_obj.theme}"
        if color_obj.type == "indexed":
            idx = color_obj.indexed
            if idx is not None and idx != 64:
                return f"indexed:{idx}"
    except Exception:
        pass
    return None


def _has_fill(fill) -> bool:
    if fill is None:
        return False
    if fill.patternType and fill.patternType != "none":
        return True
    return False


# ------------------------------------------------------------------
# Cell reading
# ------------------------------------------------------------------

def build_merge_map(ws: Worksheet) -> Dict[str, str]:
    """Return ``{coordinate: top_left_of_merge}`` for every merged cell."""
    merge_map: Dict[str, str] = {}
    for mr in ws.merged_cells.ranges:
        tl = coord(mr.min_col, mr.min_row)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                cd = coord(c, r)
                if cd != tl:
                    merge_map[cd] = tl
    return merge_map


def build_validation_map(ws: Worksheet) -> Dict[str, List[str]]:
    val_map: Dict[str, List[str]] = {}
    try:
        for dv in ws.data_validations.dataValidation:
            if dv.type == "list" and dv.formula1:
                raw = dv.formula1.strip('"')
                choices = [v.strip() for v in raw.split(",") if v.strip()]
                for cell_range in dv.sqref.ranges:
                    for r in range(cell_range.min_row, cell_range.max_row + 1):
                        for c in range(cell_range.min_col, cell_range.max_col + 1):
                            val_map[coord(c, r)] = choices
    except Exception:
        pass
    return val_map


def find_actual_used_range(ws: Worksheet) -> Tuple[int, int, int, int]:
    """Return (min_row, min_col, max_row, max_col), all 1-based."""
    dim = ws.calculate_dimension()
    if dim and dim != "A1:A1":
        try:
            parts = dim.replace("$", "").split(":")
            if len(parts) == 2:
                tl, br = parts
                tl_col = column_index_from_string("".join(c for c in tl if c.isalpha()))
                br_col = column_index_from_string("".join(c for c in br if c.isalpha()))
                tl_row = int("".join(c for c in tl if c.isdigit()))
                br_row = int("".join(c for c in br if c.isdigit()))
                if (br_row - tl_row + 1) * (br_col - tl_col + 1) <= 500_000:
                    return tl_row, tl_col, br_row, br_col
        except Exception:
            pass

    min_r = min_c = float("inf")
    max_r = max_c = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                min_r = min(min_r, cell.row)
                max_r = max(max_r, cell.row)
                min_c = min(min_c, cell.column)
                max_c = max(max_c, cell.column)
    if max_r == 0:
        return 1, 1, 1, 1
    return int(min_r), int(min_c), int(max_r), int(max_c)


def read_cell(
    cell: Cell,
    merge_map: Dict[str, str],
    val_map: Dict[str, List[str]],
    computed_values: Optional[Dict[Tuple[str, str], Any]] = None,
    sheet_name_upper: str = "",
    cached_values: Optional[Dict[Tuple[str, str], Any]] = None,
) -> CellData:
    """Read a single openpyxl Cell into a CellData DTO.

    Formula resolution order:
      1. ``computed_values`` — results from the ``formulas`` library.
      2. ``cached_values``   — Excel's own cached values (``data_only=True``).
      3. The raw formula string (last resort).
    """
    cd = coord(cell.column, cell.row)

    value = cell.value
    merged_with = merge_map.get(cd)

    formula: Optional[str] = None
    if isinstance(value, ArrayFormula):
        formula_text = getattr(value, "text", None) or ""
        formula = f"{{{formula_text}}}"
        # Try computed values first, then cached values, then formula string
        cv = (computed_values or {}).get((sheet_name_upper, cd.upper()))
        if cv is not None:
            value = cv
        else:
            cached = (cached_values or {}).get((sheet_name_upper, cd.upper()))
            value = cached if cached is not None else formula
    elif isinstance(value, str) and value.startswith("="):
        formula = value
        cv = (computed_values or {}).get((sheet_name_upper, cd.upper()))
        if cv is not None:
            value = cv
        else:
            cached = (cached_values or {}).get((sheet_name_upper, cd.upper()))
            if cached is not None:
                value = cached

    str_value: Optional[str] = str(value) if value is not None else None

    font = cell.font
    fill = cell.fill

    bg_color: Optional[str] = None
    if fill and fill.fgColor:
        bg_color = _color_hex(fill.fgColor)
    if bg_color is None and _has_fill(fill):
        bg_color = f"fill:{fill.patternType}"

    font_color: Optional[str] = None
    if font and font.color:
        font_color = _color_hex(font.color)

    return CellData(
        coordinate=cd,
        value=str_value,
        formula=formula,
        background_color=bg_color,
        font_color=font_color,
        font_size=font.size if font else None,
        font_name=font.name if font else None,
        font_bold=font.bold if font else None,
        font_italic=font.italic if font else None,
        font_underline=(
            True if font and font.underline and font.underline != "none" else None
        ),
        font_strikethrough=font.strikethrough if font else None,
        font_subscript=(
            font.vertAlign == "subscript" if font and font.vertAlign else None
        ),
        font_superscript=(
            font.vertAlign == "superscript" if font and font.vertAlign else None
        ),
        merged_with=merged_with,
        data_validation=val_map.get(cd),
    )


def read_all_cells(
    ws: Worksheet,
    computed_values: Optional[Dict[Tuple[str, str], Any]] = None,
    cached_values: Optional[Dict[Tuple[str, str], Any]] = None,
) -> Tuple[List[CellData], int, int, int, int]:
    """
    Read every cell in the used range.

    Returns (cells, min_row, min_col, max_row, max_col).
    """
    min_row, min_col, max_row, max_col = find_actual_used_range(ws)
    merge_map = build_merge_map(ws)
    val_map = build_validation_map(ws)
    sheet_name_upper = (ws.title or "").upper()

    cells: List[CellData] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cells.append(
                read_cell(
                    cell, merge_map, val_map,
                    computed_values=computed_values,
                    sheet_name_upper=sheet_name_upper,
                    cached_values=cached_values,
                )
            )
    return cells, min_row, min_col, max_row, max_col


def build_grid(
    cells: List[CellData],
) -> Dict[Tuple[int, int], CellData]:
    """Build a ``(row, col) -> CellData`` lookup from a flat cell list."""
    grid: Dict[Tuple[int, int], CellData] = {}
    for cd in cells:
        row, col = parse_coord(cd.coordinate)
        grid[(row, col)] = cd
    return grid


def slice_grid(
    grid: Dict[Tuple[int, int], CellData],
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
) -> Dict[Tuple[int, int], CellData]:
    """Return cells within the given bounding box."""
    return {
        (r, c): cd
        for (r, c), cd in grid.items()
        if min_row <= r <= max_row and min_col <= c <= max_col
    }
