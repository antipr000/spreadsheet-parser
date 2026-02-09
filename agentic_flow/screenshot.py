"""
Sheet-to-image renderer.

Strategy:
  1. Try LibreOffice headless  (most faithful rendering).
  2. Fall back to a lightweight PIL grid renderer if LibreOffice is
     not installed.

Both paths return PNG bytes.
"""

from __future__ import annotations

import logging
import os
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# LibreOffice helper
# ---------------------------------------------------------------------------

_SOFFICE_NAMES = ("soffice", "libreoffice")


def _find_soffice() -> Optional[str]:
    """Return the path to soffice / libreoffice, or None."""
    for name in _SOFFICE_NAMES:
        path = shutil.which(name)
        if path:
            return path
    # macOS typical location
    mac_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    if os.path.isfile(mac_path):
        return mac_path
    return None


def _render_with_libreoffice(
    xlsx_path: str,
    sheet_name: str,
) -> Optional[bytes]:
    """
    Convert a single sheet to PNG via LibreOffice headless.

    LibreOffice does not support exporting a single sheet directly, so
    we create a temporary copy with only the target sheet, convert that
    to PDF, then convert the PDF to PNG.
    """
    soffice = _find_soffice()
    if soffice is None:
        return None

    try:
        import openpyxl

        # Create temp workbook with only the target sheet
        src_wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        # We need to keep only the target sheet
        for sn in src_wb.sheetnames:
            if sn != sheet_name:
                del src_wb[sn]

        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_xlsx = os.path.join(tmp_dir, "sheet.xlsx")
            src_wb.save(tmp_xlsx)
            src_wb.close()

            # Convert to PNG via LibreOffice
            cmd = [
                soffice,
                "--headless",
                "--convert-to",
                "png",
                "--outdir",
                tmp_dir,
                tmp_xlsx,
            ]
            result = subprocess.run(
                cmd,
                capture_output=True,
                timeout=60,
            )
            if result.returncode != 0:
                logger.warning(
                    "LibreOffice conversion failed: %s",
                    result.stderr.decode(errors="replace")[:500],
                )
                return None

            png_path = os.path.join(tmp_dir, "sheet.png")
            if not os.path.isfile(png_path):
                # Sometimes the output name differs
                for f in Path(tmp_dir).glob("*.png"):
                    png_path = str(f)
                    break
                else:
                    logger.warning("No PNG output found from LibreOffice")
                    return None

            with open(png_path, "rb") as fh:
                return fh.read()

    except Exception:
        logger.warning(
            "LibreOffice screenshot rendering failed", exc_info=True
        )
        return None


# ---------------------------------------------------------------------------
# Lightweight PIL fallback
# ---------------------------------------------------------------------------

def _render_with_pil(
    xlsx_path: str,
    sheet_name: str,
    max_rows: int = 100,
    max_cols: int = 30,
    cell_width: int = 100,
    cell_height: int = 22,
    font_size: int = 12,
) -> Optional[bytes]:
    """
    Render a simple grid image of the sheet using PIL/Pillow.

    This is intentionally crude — it exists only as a fallback so the
    planner always gets *some* visual context.
    """
    try:
        from PIL import Image, ImageDraw, ImageFont
        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb[sheet_name]

        n_rows = min(ws.max_row or 1, max_rows)
        n_cols = min(ws.max_column or 1, max_cols)

        width = n_cols * cell_width + 1
        height = n_rows * cell_height + 1

        img = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(img)

        # Try to get a monospace font; fall back to default
        try:
            font = ImageFont.truetype("DejaVuSansMono.ttf", font_size)
        except Exception:
            try:
                font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf", font_size)
            except Exception:
                font = ImageFont.load_default()

        for r in range(1, n_rows + 1):
            for c in range(1, n_cols + 1):
                x0 = (c - 1) * cell_width
                y0 = (r - 1) * cell_height
                x1 = x0 + cell_width
                y1 = y0 + cell_height

                cell = ws.cell(row=r, column=c)

                # Background for bold cells
                if cell.font and cell.font.bold:
                    draw.rectangle([x0, y0, x1, y1], fill="#E0E0E0")

                # Grid lines
                draw.rectangle([x0, y0, x1, y1], outline="#CCCCCC")

                # Cell text
                val = cell.value
                if val is not None:
                    text = str(val)[:12]  # truncate for display
                    draw.text(
                        (x0 + 2, y0 + 2),
                        text,
                        fill="black",
                        font=font,
                    )

        wb.close()

        import io
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()

    except ImportError:
        logger.warning("Pillow not installed — cannot render PIL fallback")
        return None
    except Exception:
        logger.warning("PIL screenshot rendering failed", exc_info=True)
        return None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def render_sheet_screenshot(
    xlsx_path: str,
    sheet_name: str,
) -> Optional[bytes]:
    """
    Return PNG bytes of the rendered sheet, or None if rendering failed.

    Tries LibreOffice first, falls back to PIL.
    """
    png = _render_with_libreoffice(xlsx_path, sheet_name)
    if png is not None:
        logger.info("  Screenshot rendered via LibreOffice (%d bytes)", len(png))
        return png

    png = _render_with_pil(xlsx_path, sheet_name)
    if png is not None:
        logger.info("  Screenshot rendered via PIL fallback (%d bytes)", len(png))
        return png

    logger.warning("  Could not render screenshot for sheet '%s'", sheet_name)
    return None
