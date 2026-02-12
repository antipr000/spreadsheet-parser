"""
Spreadsheet parser — CLI entry point.

Usage:
    python parser.py <excel_file> [--output <output.json>] [--sheet <sheet_name>]

Loads an Excel workbook, extracts all semantic blocks from every sheet
(headings, tables, key-value regions, text, charts), groups them into
chunks, and writes the result as a single JSON file.

If --sheet is provided, only that worksheet is processed.
"""

from __future__ import annotations

import argparse
import logging
import os
import re
import sys
from pathlib import Path
from typing import Any, Dict, Optional, Tuple

import dotenv
import formulas
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter

from dto.blocks import TableBlock, Block
from dto.output import SheetResult, WorkbookResult
from extractors.sheet import SheetExtractor
from grouping import group_blocks_into_chunks
from utils.html import render_table_html
from utils.row_groups import detect_row_groups

dotenv.load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(name)s  %(message)s",
)
logger = logging.getLogger(__name__)


# -------------------------------------------------------------------
# Post-processing: fill in HTML for TableBlocks
# -------------------------------------------------------------------


def _enrich_blocks(blocks: list[Block]) -> list[Block]:
    """
    Apply post-processing enrichment to blocks before serialisation.
      - Render HTML for table blocks
      - Detect row groupings within table blocks
    """
    for block in blocks:
        if isinstance(block, TableBlock):
            if not block.html:
                block.html = render_table_html(
                    heading=block.heading,
                    data=block.data,
                    footer=block.footer,
                )
            # Detect hierarchical row groupings (only for tables that
            # look like they might have groups; no-op for flat tables).

            # This is not working as expected, we need to visit this later.
            # if not block.row_groups:
            #     try:
            #         detect_row_groups(block)
            #     except Exception:
            #         logger.warning(
            #             "Row group detection failed for table %s",
            #             block.bounding_box.top_left,
            #             exc_info=True,
            #         )
    return blocks


# -------------------------------------------------------------------
# Main pipeline
# -------------------------------------------------------------------


def _compute_formula_values(file_path: str) -> Dict[Tuple[str, str], Any]:
    """
    Use the ``formulas`` library to evaluate every formula in the workbook
    and return a lookup:  ``(sheet_name_upper, cell_coordinate) → value``.

    Sheet names are normalised to uppercase for case-insensitive matching.
    """
    computed: Dict[Tuple[str, str], Any] = {}
    try:
        xl_model = formulas.ExcelModel().loads(file_path).finish()
        results = xl_model.calculate()

        # results keys look like  "'[file.xlsx]SHEET NAME'!E2"  or range variants
        file_stem = Path(file_path).name
        pattern = re.compile(
            r"'\[" + re.escape(file_stem) + r"\](.+?)'!([A-Z]+\d+)$",
            re.IGNORECASE,
        )
        for key, val in results.items():
            m = pattern.match(str(key))
            if not m:
                continue
            sheet = m.group(1).upper()
            coord = m.group(2).upper()

            # Unwrap numpy scalars / 1-element arrays
            v = val
            if hasattr(v, "value"):
                # Ranges object — get the underlying array
                v = getattr(v, "value", v)
            if isinstance(v, np.ndarray):
                if v.size == 1:
                    v = v.flat[0]
                else:
                    continue  # skip multi-cell range results
            if isinstance(v, (np.integer, np.floating)):
                v = v.item()

            computed[(sheet, coord)] = v
    except Exception:
        logger.warning(
            "Formula evaluation failed — computed values will be unavailable",
            exc_info=True,
        )

    return computed


def _load_cached_values(file_path: str) -> Dict[Tuple[str, str], Any]:
    """
    Open the workbook with ``data_only=True`` to read Excel's own cached
    formula results.  Returns a lookup ``(SHEET_NAME_UPPER, COORD) → value``.

    Fallback for when the ``formulas`` library cannot compute a value.
    """
    cached: Dict[Tuple[str, str], Any] = {}
    try:
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        for ws_name in wb_data.sheetnames:
            ws = wb_data[ws_name]
            sheet_upper = ws_name.upper()
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if v is None:
                        continue
                    if isinstance(v, str) and v.startswith("="):
                        continue
                    coord_str = f"{get_column_letter(cell.column)}{cell.row}"
                    cached[(sheet_upper, coord_str)] = v
        wb_data.close()
    except Exception:
        logger.warning(
            "Failed to load cached formula values (data_only workbook)",
            exc_info=True,
        )
    return cached


def parse_workbook(
    file_path: str,
    sheet_name_filter: Optional[str] = None,
) -> WorkbookResult:
    """
    Parse an Excel workbook and return a structured ``WorkbookResult``.

    If *sheet_name_filter* is provided, only that worksheet is processed.
    """
    logger.info("Loading workbook: %s", file_path)

    workbook = openpyxl.load_workbook(
        file_path,
        data_only=False,
        read_only=False,
        keep_links=True,
        keep_vba=True,
        rich_text=True,
    )

    if sheet_name_filter and sheet_name_filter not in workbook.sheetnames:
        logger.error(
            "Worksheet '%s' not found. Available sheets: %s",
            sheet_name_filter,
            workbook.sheetnames,
        )
        raise ValueError(
            f"Worksheet '{sheet_name_filter}' not found in workbook"
        )

    # Compute formula values up-front
    logger.info("Computing formula values...")
    computed_values = _compute_formula_values(file_path)
    logger.info("  -> %d formula value(s) computed", len(computed_values))

    # Load Excel's cached formula results as a fallback
    logger.info("Loading cached formula values (data_only)...")
    cached_values = _load_cached_values(file_path)
    logger.info("  -> %d cached value(s) loaded", len(cached_values))

    extractor = SheetExtractor()
    sheet_results: list[SheetResult] = []

    sheets_to_process = (
        [sheet_name_filter] if sheet_name_filter else workbook.sheetnames
    )

    for sheet_name in sheets_to_process:
        logger.info("Processing sheet: %s", sheet_name)
        ws = workbook[sheet_name]

        try:
            # Extract all blocks
            blocks = extractor.extract(
                ws, workbook,
                computed_values=computed_values,
                cached_values=cached_values,
            )

            # Enrich (e.g. render HTML for tables)
            blocks = _enrich_blocks(blocks)

            # Group headings with their content blocks into chunks
            chunks = group_blocks_into_chunks(blocks)

            sheet_results.append(
                SheetResult(
                    sheet_name=sheet_name,
                    chunks=chunks,
                )
            )
            logger.info(
                "  -> %d block(s) in %d chunk(s)",
                sum(len(v) for v in chunks.values()),
                len(chunks),
            )
        except Exception:
            logger.exception(
                "Failed to process sheet '%s' — adding empty result", sheet_name
            )
            sheet_results.append(SheetResult(sheet_name=sheet_name))

    return WorkbookResult(
        file_name=Path(file_path).name,
        sheets=sheet_results,
    )


# -------------------------------------------------------------------
# CLI
# -------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Parse an Excel workbook into a structured JSON document.",
    )
    parser.add_argument(
        "excel_file",
        help="Path to the .xlsx file to parse",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Output JSON file path (default: <input_name>_output.json)",
    )
    parser.add_argument(
        "-s",
        "--sheet",
        default=None,
        help="Name of a single worksheet to process (default: all sheets)",
    )
    args = parser.parse_args()

    excel_path = args.excel_file
    if not os.path.isfile(excel_path):
        logger.error("File not found: %s", excel_path)
        sys.exit(1)

    # Determine output path
    if args.output:
        output_path = args.output
    else:
        stem = Path(excel_path).stem
        output_path = f"{stem}_output.json"

    # Run pipeline
    result = parse_workbook(excel_path, sheet_name_filter=args.sheet)

    # Serialize to JSON
    json_str = result.model_dump_json(indent=2, exclude_none=True)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(json_str)

    logger.info("Output written to %s", output_path)


if __name__ == "__main__":
    main()
