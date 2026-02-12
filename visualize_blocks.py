"""
Block Visualizer — creates a new workbook containing only the parsed
sheet(s), with cells colored by their detected block.

Usage:
    python visualize_blocks.py <parser_output.json> [--workbook-dir <dir>] [-o <output.xlsx>]
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
)
logger = logging.getLogger(__name__)

# ── Distinct pastel colors (ARGB hex, no leading #) ─────────────────
_BLOCK_COLORS = [
    "FFB3E5FC",  # light blue
    "FFC8E6C9",  # light green
    "FFFFF9C4",  # light yellow
    "FFFFCCBC",  # light orange
    "FFE1BEE7",  # light purple
    "FFB2DFDB",  # light teal
    "FFF8BBD0",  # light pink
    "FFD7CCC8",  # light brown
    "FFBBDEFB",  # blue 100
    "FFDCEDC8",  # green 100
    "FFFFF176",  # yellow 300
    "FFFFAB91",  # deep orange 200
    "FFCE93D8",  # purple 200
    "FF80CBC4",  # teal 200
    "FFF48FB1",  # pink 200
    "FFBCAAA4",  # brown 200
    "FF90CAF9",  # blue 200
    "FFA5D6A7",  # green 200
    "FFFFD54F",  # amber 300
    "FFEF9A9A",  # red 200
    "FFB0BEC5",  # blue grey 200
    "FFFFE082",  # amber 200
    "FF80DEEA",  # cyan 200
    "FFFF8A65",  # deep orange 300
    "FFAED581",  # light green 300
]

_BORDER_COLORS = [
    "FF0288D1",  # blue
    "FF388E3C",  # green
    "FFF9A825",  # yellow
    "FFE64A19",  # orange
    "FF7B1FA2",  # purple
    "FF00796B",  # teal
    "FFC2185B",  # pink
    "FF5D4037",  # brown
    "FF1565C0",  # blue 800
    "FF558B2F",  # green 800
    "FFF57F17",  # yellow 900
    "FFBF360C",  # deep orange 900
    "FF6A1B9A",  # purple 800
    "FF00695C",  # teal 800
    "FFAD1457",  # pink 800
    "FF4E342E",  # brown 800
    "FF0D47A1",  # blue 900
    "FF2E7D32",  # green 800
    "FFFF8F00",  # amber 800
    "FFC62828",  # red 800
    "FF455A64",  # blue grey 700
    "FFFF8F00",  # amber 800
    "FF00838F",  # cyan 800
    "FFD84315",  # deep orange 800
    "FF33691E",  # light green 900
]

_LABEL_FONT = Font(bold=True, size=9, color="FF000000")
_BLOCK_LABEL_FONT = Font(bold=True, size=8, color="FFFFFFFF")


# ── Helpers ──────────────────────────────────────────────────────────


def _parse_coord(coord: str):
    """Parse 'E3' → (row=3, col=5)."""
    col_str = "".join(c for c in coord if c.isalpha())
    row_num = int("".join(c for c in coord if c.isdigit()) or "0")
    col_num = column_index_from_string(col_str) if col_str else 1
    return row_num, col_num


def _make_border(color: str) -> Border:
    """Create a thin border with the given ARGB color on all four sides."""
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _copy_sheet(src_wb, dst_wb, sheet_name: str) -> None:
    """
    Copy a sheet from *src_wb* into *dst_wb*, preserving cell values,
    formatting, merged cells, column widths, and row heights.
    """
    src_ws = src_wb[sheet_name]
    dst_ws = dst_wb.create_sheet(title=sheet_name)

    # Copy merged cell ranges
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    # Copy column dimensions
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
        dst_ws.column_dimensions[col_letter].hidden = dim.hidden

    # Copy row dimensions
    for row_num, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = dim.height
        dst_ws.row_dimensions[row_num].hidden = dim.hidden

    # Copy cells (value + style)
    for row in src_ws.iter_rows():
        for src_cell in row:
            dst_cell = dst_ws.cell(
                row=src_cell.row,
                column=src_cell.column,
                value=src_cell.value,
            )
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy(src_cell.protection)
                dst_cell.alignment = copy(src_cell.alignment)

    # Copy freeze panes
    dst_ws.freeze_panes = src_ws.freeze_panes


# ── Core logic ───────────────────────────────────────────────────────


def visualize(json_path: str, workbook_dir: str, output_path: str) -> None:
    """
    Read the parser JSON, copy only the parsed sheet(s) into a new
    workbook, color each block, add labels, and save.
    """
    # Load parser output
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    file_name = data.get("file_name", "")
    if not file_name:
        logger.error("JSON has no 'file_name' field")
        sys.exit(1)

    workbook_path = Path(workbook_dir) / file_name
    if not workbook_path.is_file():
        logger.error("Workbook not found: %s", workbook_path)
        sys.exit(1)

    logger.info("Loading source workbook: %s", workbook_path)
    src_wb = openpyxl.load_workbook(str(workbook_path))

    sheets_data = data.get("sheets", [])
    if not sheets_data:
        logger.warning("JSON contains no sheets")
        return

    # Create a fresh output workbook (remove the default empty sheet later)
    dst_wb = openpyxl.Workbook()
    default_sheet = dst_wb.active

    # Track legend entries across all sheets
    legend_entries = []

    for sheet_data in sheets_data:
        sheet_name = sheet_data.get("sheet_name", "")
        if sheet_name not in src_wb.sheetnames:
            logger.warning("Sheet '%s' not found in workbook — skipping", sheet_name)
            continue

        chunks = sheet_data.get("chunks", {})
        if not chunks:
            logger.info("Sheet '%s' has no chunks — skipping", sheet_name)
            continue

        # Copy the sheet into the new workbook
        logger.info("Copying sheet '%s' into output workbook", sheet_name)
        _copy_sheet(src_wb, dst_wb, sheet_name)
        ws = dst_wb[sheet_name]

        logger.info("Processing sheet '%s': %d chunk(s)", sheet_name, len(chunks))

        # Sort chunk keys numerically (block0, block1, ...)
        sorted_keys = sorted(
            chunks.keys(),
            key=lambda k: int("".join(c for c in k if c.isdigit()) or "0"),
        )

        color_idx = len(legend_entries)  # continue color sequence across sheets
        for chunk_key in sorted_keys:
            blocks = chunks[chunk_key]

            for block_i, block in enumerate(blocks):
                block_type = block.get("block_type", "unknown")
                bbox = block.get("bounding_box", {})
                tl_str = bbox.get("top_left", "")
                br_str = bbox.get("bottom_right", "")

                if not tl_str or not br_str:
                    logger.warning(
                        "  %s block %d has no bounding box — skipping",
                        chunk_key,
                        block_i,
                    )
                    continue

                tl_row, tl_col = _parse_coord(tl_str)
                br_row, br_col = _parse_coord(br_str)

                # Pick color for this block
                fill_argb = _BLOCK_COLORS[color_idx % len(_BLOCK_COLORS)]
                border_argb = _BORDER_COLORS[color_idx % len(_BORDER_COLORS)]
                fill = PatternFill(
                    start_color=fill_argb,
                    end_color=fill_argb,
                    fill_type="solid",
                )
                border = _make_border(border_argb)

                # Color every cell in the bounding box
                for r in range(tl_row, br_row + 1):
                    for c in range(tl_col, br_col + 1):
                        cell = ws.cell(row=r, column=c)
                        if isinstance(cell, MergedCell):
                            continue  # skip non-master merged cells
                        cell.fill = fill
                        cell.border = border

                # Prepend a visible [block_type] label to the top-left cell
                tl_cell = ws.cell(row=tl_row, column=tl_col)
                if not isinstance(tl_cell, MergedCell):
                    existing_val = tl_cell.value
                    if existing_val is not None:
                        tl_cell.value = f"[{block_type}] {existing_val}"
                    else:
                        tl_cell.value = f"[{block_type}]"
                    tl_cell.font = Font(
                        bold=True,
                        size=tl_cell.font.size or 10,
                        color=border_argb,
                    )
                    # Also add a comment with full details
                    comment_text = f"{chunk_key} | {block_type}\n{tl_str}:{br_str}"
                    tl_cell.comment = Comment(comment_text, "block-visualizer")

                logger.info(
                    "  %s [%d]: %s  %s:%s  → color #%d",
                    chunk_key,
                    block_i,
                    block_type,
                    tl_str,
                    br_str,
                    color_idx,
                )

                legend_entries.append(
                    {
                        "chunk_key": chunk_key,
                        "block_type": block_type,
                        "bbox": f"{tl_str}:{br_str}",
                        "sheet": sheet_name,
                        "fill_argb": fill_argb,
                    }
                )

                color_idx += 1

    # Remove the default empty sheet created by openpyxl
    if default_sheet is not None and default_sheet.title not in [
        sd.get("sheet_name", "") for sd in sheets_data
    ]:
        dst_wb.remove(default_sheet)

    # ── Add a legend sheet ───────────────────────────────────────────
    ws_legend = dst_wb.create_sheet("_Block Legend")

    headers = ["Block ID", "Type", "Bounding Box", "Sheet", "Color"]
    col_widths = [14, 14, 18, 18, 10]
    for i, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws_legend.cell(row=1, column=i, value=header)
        cell.font = _LABEL_FONT
        col_letter = get_column_letter(i)
        ws_legend.column_dimensions[col_letter].width = width

    for row_idx, entry in enumerate(legend_entries, start=2):
        ws_legend.cell(row=row_idx, column=1, value=entry["chunk_key"])
        ws_legend.cell(row=row_idx, column=2, value=entry["block_type"])
        ws_legend.cell(row=row_idx, column=3, value=entry["bbox"])
        ws_legend.cell(row=row_idx, column=4, value=entry["sheet"])
        color_cell = ws_legend.cell(row=row_idx, column=5, value="")
        color_cell.fill = PatternFill(
            start_color=entry["fill_argb"],
            end_color=entry["fill_argb"],
            fill_type="solid",
        )

    dst_wb.save(output_path)
    logger.info("Annotated workbook saved to: %s", output_path)


# ── CLI ──────────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Visualize detected blocks by coloring cells in a new workbook "
            "containing only the parsed sheet(s).  The source workbook path "
            "is read from the JSON's 'file_name' field."
        ),
    )
    parser.add_argument(
        "json_file",
        help="Path to the parser output JSON",
    )
    parser.add_argument(
        "--workbook-dir",
        default=".",
        help=(
            "Directory containing the original .xlsx workbook "
            "(default: current directory)"
        ),
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Output .xlsx path (default: <workbook>_visualized.xlsx)",
    )
    args = parser.parse_args()

    # Derive output path
    if args.output:
        output_path = args.output
    else:
        with open(args.json_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        file_name = data.get("file_name", "output.xlsx")
        stem = Path(file_name).stem
        output_path = f"{stem}_visualized.xlsx"

    visualize(args.json_file, args.workbook_dir, output_path)


if __name__ == "__main__":
    main()
