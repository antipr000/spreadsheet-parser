"""
Detector for table regions.

Heuristic rules:
  - Region has ≥ 3 rows and ≥ 2 columns
  - At least one row looks like a header (bold / coloured / label-like)
  - Body rows have a repeating columnar pattern (most cells occupied)
  - This is the **default / fallback** detector — if no other type matched
    and the region has enough structure, it's probably a table.

The AI path reuses the structured table-schema prompt and returns full
header / body / footer decomposition.
"""

from __future__ import annotations

import logging
from collections import Counter
from typing import List, Optional

from openpyxl.utils import column_index_from_string, get_column_letter

from ai.factory import get_decision_service
from ai.response_parser import parse_llm_json
from detection.base import Detector
from dto.ai import TableSchemaDTO
from dto.blocks import Block, TableBlock
from dto.cell_data import CellData
from dto.coordinate import BoundingBox
from dto.region import RegionData
from prompts.detection import get_table_detection_prompt

logger = logging.getLogger(__name__)


class TableDetector(Detector):

    # ------------------------------------------------------------------
    # Heuristic helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _looks_numeric(val: str) -> bool:
        """Return True if the string looks like a number."""
        cleaned = val.strip().lstrip("$€£").replace(",", "").replace("%", "")
        try:
            float(cleaned)
            return True
        except ValueError:
            return False

    @staticmethod
    def _cell_type(cell: CellData) -> str:
        """Classify a cell's value into a rough type bucket."""
        if cell.value is None:
            return "empty"
        if cell.formula:
            return "formula"
        if TableDetector._looks_numeric(cell.value):
            return "numeric"
        return "text"

    def _first_row_looks_like_header(
        self, region: RegionData, all_columns: List[int]
    ) -> bool:
        """
        Content-based check: does the first row look like a header row
        rather than a data row?

        Returns True if the first row should be treated as a header.
        Returns False if it looks like data (table has no header).

        Signals that the first row IS a header:
          - First row is all text while at least one column has predominantly
            numeric body values below it.
          - First row values are shorter / fewer words than body values in the
            same column on average (labels vs data instances).

        Signals that the first row is NOT a header (= data):
          - First row has the same type profile as subsequent rows (e.g. all
            rows are text+numeric in the same columns).
        """
        first_row = region.min_row
        body_start = region.min_row + 1

        if body_start > region.max_row:
            return True  # Only one row — treat as header by default

        # Build type profile for first row and body rows per column
        first_row_types: List[str] = []
        body_type_profiles: List[List[str]] = []

        for col in all_columns:
            fc = region.cell_at(first_row, col)
            first_row_types.append(self._cell_type(fc) if fc else "empty")

            col_types = []
            for r in range(body_start, region.max_row + 1):
                bc = region.cell_at(r, col)
                if bc and bc.value is not None:
                    col_types.append(self._cell_type(bc))
            body_type_profiles.append(col_types)

        # Check 1: If a column's first-row type is "text" but the body is
        # predominantly "numeric", the first row is a header label.
        for i, col in enumerate(all_columns):
            if first_row_types[i] != "text":
                continue
            body_types = body_type_profiles[i]
            if not body_types:
                continue
            numeric_ratio = body_types.count("numeric") / len(body_types)
            if numeric_ratio >= 0.6:
                return True  # text header above numeric data

        # Check 2: If the first row's type profile matches the majority
        # of body rows' profiles, it's data, not a header.
        # Compare per-column: for each column, does the first row's type
        # match what most body rows have?
        matches = 0
        for i, col in enumerate(all_columns):
            body_types = body_type_profiles[i]
            if not body_types:
                continue
            # Most common type in body for this column
            most_common = Counter(body_types).most_common(1)[0][0]
            if first_row_types[i] == most_common:
                matches += 1

        if len(all_columns) > 0 and matches >= len(all_columns) * 0.8:
            return False  # first row looks like data

        # Default: treat first row as header (safer assumption)
        return True

    # ------------------------------------------------------------------
    # Heuristic
    # ------------------------------------------------------------------

    def detect(self, region: RegionData) -> Optional[Block]:
        # Need enough structure to be a table
        if region.num_rows < 3 or region.num_cols < 2:
            return None

        non_empty = region.non_empty_cells
        if len(non_empty) < 4:
            return None

        # Check that data is spread across multiple columns
        occupied_cols: set = set()
        for c in non_empty:
            col_str = "".join(ch for ch in c.coordinate if ch.isalpha())
            occupied_cols.add(column_index_from_string(col_str))

        if len(occupied_cols) < 2:
            return None

        # Try to identify header row(s): bold or background-coloured
        header_rows: List[int] = []
        body_rows: List[int] = []

        for r in range(region.min_row, region.max_row + 1):
            row_cells = [
                cd
                for cd in (region.cell_at(r, c) for c in range(region.min_col, region.max_col + 1))
                if cd is not None and cd.value is not None
            ]
            if not row_cells:
                continue

            is_header_row = (
                all(c.font_bold for c in row_cells)
                or all(c.background_color is not None for c in row_cells)
            )
            if is_header_row and not body_rows:
                # Only rows before the first body row can be headers
                header_rows.append(r)
            else:
                body_rows.append(r)

        # A table without any body rows isn't a table
        if not body_rows:
            return None

        all_columns = sorted(occupied_cols)

        # If no formatting-based headers detected, check whether the first
        # row looks like a header by content (type analysis).
        if not header_rows:
            if self._first_row_looks_like_header(region, all_columns):
                header_rows = [body_rows.pop(0)]
                if not body_rows:
                    return None
            # else: genuinely headerless table — header_rows stays empty

        # Collect cells into heading / data lists
        heading_cells = self._collect_cells(region, header_rows, all_columns)
        data_cells = self._collect_cells(region, body_rows, all_columns)

        all_cells = heading_cells + data_cells
        return TableBlock(
            bounding_box=region.bounding_box,
            heading=heading_cells,
            data=data_cells,
            footer=[],
            cells=all_cells,
        )

    @staticmethod
    def _collect_cells(
        region: RegionData, rows: List[int], cols: List[int]
    ) -> List[CellData]:
        out: List[CellData] = []
        for r in rows:
            for c in cols:
                cd = region.cell_at(r, c)
                if cd:
                    out.append(cd)
        return out

    # ------------------------------------------------------------------
    # AI-assisted
    # ------------------------------------------------------------------

    def detect_with_ai(self, region: RegionData) -> Optional[Block]:
        prompt = get_table_detection_prompt(region.non_empty_cells)
        ai = get_decision_service()
        raw = ai.get_decision(prompt)

        parsed = parse_llm_json(raw)
        if not isinstance(parsed, dict):
            return None

        if not parsed.get("is_table", False):
            return None

        raw_tables = parsed.get("tables", [])
        if not raw_tables:
            return None

        # For now, return the first table the LLM finds in this region.
        # (The region was already split by heuristics, so usually there's one.)
        blocks: List[TableBlock] = []
        for item in raw_tables:
            try:
                schema = TableSchemaDTO.model_validate(item)
            except Exception as exc:
                logger.warning("Skipping invalid table schema from LLM: %s", exc)
                continue

            if not self._validate_schema(schema, region):
                logger.warning("Discarding out-of-bounds table schema: %s", schema)
                continue

            blocks.append(self._schema_to_block(schema, region))

        return blocks[0] if blocks else None

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _validate_schema(schema: TableSchemaDTO, region: RegionData) -> bool:
        """Sanity-check a TableSchemaDTO against the region bounds."""
        try:
            tl_col = column_index_from_string(
                "".join(c for c in schema.top_left if c.isalpha())
            )
            tl_row = int("".join(c for c in schema.top_left if c.isdigit()))
            br_col = column_index_from_string(
                "".join(c for c in schema.bottom_right if c.isalpha())
            )
            br_row = int("".join(c for c in schema.bottom_right if c.isdigit()))
        except Exception:
            return False

        if tl_row > br_row or tl_col > br_col:
            return False

        # Check rough overlap with the region
        if br_row < region.min_row or tl_row > region.max_row:
            return False
        if br_col < region.min_col or tl_col > region.max_col:
            return False

        # No row overlap between sections
        all_rows = set(schema.header_rows) | set(schema.body_rows) | set(schema.footer_rows)
        total = len(schema.header_rows) + len(schema.body_rows) + len(schema.footer_rows)
        if len(all_rows) < total:
            return False

        return True

    @staticmethod
    def _schema_to_block(schema: TableSchemaDTO, region: RegionData) -> TableBlock:
        """Convert a validated TableSchemaDTO into a TableBlock using the region grid."""
        header_col_idx = [column_index_from_string(c) for c in schema.header_columns]
        body_col_idx = [column_index_from_string(c) for c in schema.body_columns]
        footer_col_idx = [column_index_from_string(c) for c in schema.footer_columns]

        def _gather(rows: List[int], cols: List[int]) -> List[CellData]:
            out: List[CellData] = []
            for r in rows:
                for c in cols:
                    cd = region.cell_at(r, c)
                    if cd:
                        out.append(cd)
            return out

        heading = _gather(schema.header_rows, header_col_idx)
        data = _gather(schema.body_rows, body_col_idx)
        footer = _gather(schema.footer_rows, footer_col_idx)
        all_cells = heading + data + footer

        return TableBlock(
            bounding_box=BoundingBox(
                top_left=schema.top_left,
                bottom_right=schema.bottom_right,
            ),
            heading=heading,
            data=data,
            footer=footer,
            cells=all_cells,
        )
